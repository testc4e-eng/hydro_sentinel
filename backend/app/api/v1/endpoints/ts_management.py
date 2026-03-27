
from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from app.db.session import get_db
from typing import Optional, List, Any, Dict, Tuple, Set
from datetime import datetime, timezone
from statistics import median
from pydantic import BaseModel
from uuid import UUID
import tempfile
import os
import shutil
import csv
import sys
import re
import unicodedata
import time
import gc

router = APIRouter()

print("\n*** LOADING TS_MANAGEMENT v_FIX_INJECT_V2 ***\n")


TIMESTAMP_KEYWORDS = ['timestamp', 'time', 'date', 'datetime', 'horodatage', 'date/heure']

SOURCE_HINT_TOKENS: Dict[str, str] = {
    "obs": "OBS",
    "observation": "OBS",
    "observations": "OBS",
    "observe": "OBS",
    "observee": "OBS",
    "observees": "OBS",
    "history": "OBS",
    "historique": "OBS",
    "sim": "SIM",
    "simu": "SIM",
    "simule": "SIM",
    "simulee": "SIM",
    "simulees": "SIM",
    "simulated": "SIM",
    "simulation": "SIM",
    "arome": "AROME",
    "arom": "AROME",
    "ecmwf": "ECMWF",
    "ecmw": "ECMWF",
}

SOURCE_DECORATION_TOKENS = {
    "prev",
    "previ",
    "prevision",
    "previsions",
    "forecast",
    "model",
    "modele",
    "source",
}

PRECIP_HINT_TOKENS = {
    "precip",
    "precipitation",
    "precipitations",
    "pluie",
    "rain",
    "rainfall",
    "neige",
    "snow",
}

HYDRO_HINT_TOKENS = {
    "flow",
    "debit",
    "inflow",
    "apport",
    "apports",
    "solide",
    "solid",
    "sediment",
    "volume",
    "cote",
    "lacher",
    "lachers",
}

VARIABLE_COLUMN_ALIASES = {
    "flow_m3s": "debit_m3s",
    "debit_m3s": "flow_m3s",
    "inflow_m3s": "apport_m3s",
    "apport_m3s": "inflow_m3s",
    "precip_mm": "pluie_mm",
    "pluie_mm": "precip_mm",
    "precipitation_mm": "precip_mm",
    "precipitations_mm": "precip_mm",
    "flow": "flow_m3s",
    "debit": "flow_m3s",
    "inflow": "inflow_m3s",
    "apport": "inflow_m3s",
    "pluie": "precip_mm",
    "precip": "precip_mm",
    "volume": "volume_hm3",
}

BARRAGE_VARIABLE_CODES = {'lacher_m3s', 'volume_k', 'cote_m', 'lachers', 'volume'}
QUALITY_FLAG_COLUMN_TOKENS = {
    "quality_flag",
    "qc_flag",
    "quality",
    "qualite",
    "qc",
    "flag_qualite",
    "flagqualite",
    "flag",
}

KNOWN_SOURCE_CODES: Tuple[str, ...] = ("OBS", "SIM", "AROME", "ECMWF")

STATION_DECORATION_TOKENS = {
    "debit", "flow", "inflow", "apport", "apports", "volume", "cote", "lacher", "lachers",
    "pluie", "precip", "precipitation", "precipitations", "rain", "rainfall",
    "mm", "m3", "m3s", "m3h", "hm3", "cms", "m", "s", "hr", "h", "1hr", "24h",
    "station", "st", "poste", "hydro", "hydrologique",
}


def _get_pandas():
    try:
        import pandas as pd  # type: ignore
        return pd
    except Exception:
        return None


def _is_empty_cell(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _normalize_headers(raw_headers: List[Any]) -> List[str]:
    headers: List[str] = []
    seen: Dict[str, int] = {}
    for idx, value in enumerate(raw_headers):
        base = str(value).strip() if value is not None else ""
        if not base:
            base = f"unnamed_{idx + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers


def _find_header_row_index(preview_rows: List[List[Any]]) -> int:
    for i, row in enumerate(preview_rows[:10]):
        non_empty = [v for v in row if not _is_empty_cell(v)]
        if len(non_empty) < 2:
            continue

        strict_match = any(str(v).lower().strip() in TIMESTAMP_KEYWORDS for v in non_empty)
        if strict_match:
            return i

        relaxed_match = any(
            any(k in str(v).lower().strip() for k in TIMESTAMP_KEYWORDS) and len(str(v).strip()) < 20
            for v in non_empty
        )
        if relaxed_match:
            return i

    return 0


def _load_with_pandas(file_path: str, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    pd = _get_pandas()
    if pd is None:
        raise RuntimeError("pandas_unavailable")

    if filename.lower().endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df_preview = pd.read_excel(file_path, header=None, nrows=10)
        preview_rows = [list(row.values) for _, row in df_preview.iterrows()]
        header_row_idx = _find_header_row_index(preview_rows)
        df = pd.read_excel(file_path, header=header_row_idx)

    df.columns = [str(c).strip() for c in df.columns]
    columns = [str(c).strip() for c in df.columns]
    records = df.to_dict(orient='records')
    return columns, records


def _load_with_builtin_parsers(file_path: str, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    lower = filename.lower()

    raw_rows: List[List[Any]] = []
    if lower.endswith('.csv'):
        with open(file_path, 'r', encoding='utf-8-sig', errors='replace', newline='') as f:
            reader = csv.reader(f)
            raw_rows = [row for row in reader]
    else:
        def _resolve_openpyxl_loader():
            try:
                from openpyxl import load_workbook  # type: ignore
                return load_workbook
            except Exception:
                pass

            fallback_paths = []
            conda_prefix = os.environ.get("CONDA_PREFIX")
            if conda_prefix:
                fallback_paths.append(os.path.join(conda_prefix, "Lib", "site-packages"))
            if os.name == "nt":
                fallback_paths.append(r"C:\anaconda\Lib\site-packages")
                fallback_paths.append(r"C:\ProgramData\anaconda3\Lib\site-packages")

            for path in fallback_paths:
                if path and os.path.isdir(path) and path not in sys.path:
                    sys.path.append(path)
                try:
                    from openpyxl import load_workbook  # type: ignore
                    return load_workbook
                except Exception:
                    continue

            raise HTTPException(
                status_code=503,
                detail="Neither pandas nor openpyxl is available to parse file.",
            )

        load_workbook = _resolve_openpyxl_loader()

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))
        wb.close()

    if not raw_rows:
        return [], []

    header_idx = _find_header_row_index(raw_rows)
    headers = _normalize_headers(raw_rows[header_idx])

    records: List[Dict[str, Any]] = []
    for raw in raw_rows[header_idx + 1:]:
        if all(_is_empty_cell(v) for v in raw):
            continue
        row_dict: Dict[str, Any] = {}
        for i, col in enumerate(headers):
            row_dict[col] = raw[i] if i < len(raw) else None
        records.append(row_dict)

    return headers, records


def _load_tabular_data(file_path: str, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        return _load_with_pandas(file_path, filename)
    except Exception as exc:
        print(f"Warning: pandas parser unavailable/failed ({exc}), falling back to builtin parser.")
        return _load_with_builtin_parsers(file_path, filename)


def _load_template_entity_alias_map(file_path: str, filename: str) -> Dict[str, str]:
    """
    Extract optional alias map from template sheet ('Bassins'/'Stations'):
    Code -> Name. Keys are normalized with _entity_lookup_keys.
    """
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return {}

    alias_map: Dict[str, str] = {}

    def _register(code_value: Any, name_value: Any, external_value: Any = None) -> None:
        code_raw = str(code_value or "").strip()
        name_raw = str(name_value or "").strip()
        if not code_raw or not name_raw:
            return
        for key in _alias_lookup_keys(code_raw):
            alias_map.setdefault(key, name_raw)
        # Directly map external source column names (e.g. "AIN AICHA_Pluie 1hr (mm)")
        # to canonical DB station names.
        external_raw = str(external_value or "").strip()
        if external_raw:
            for key in _alias_lookup_keys(external_raw):
                alias_map.setdefault(key, name_raw)

    pd = _get_pandas()
    if pd is not None:
        try:
            xls = pd.ExcelFile(file_path)
            try:
                sheet_name = "Bassins" if "Bassins" in xls.sheet_names else ("Stations" if "Stations" in xls.sheet_names else None)
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    if not df.empty:
                        columns = [str(c).strip().lower() for c in df.columns]
                        code_idx = 0
                        name_idx = 1 if len(columns) > 1 else 0
                        ext_idx: Optional[int] = None
                        for idx, col_name in enumerate(columns):
                            if col_name in {"code", "id", "bassin_code", "station_code"}:
                                code_idx = idx
                            if col_name in {"nom", "name", "bassin_nom", "station_nom", "nom station"}:
                                name_idx = idx
                            if (
                                "nom station fichier" in col_name
                                or "station fichier" in col_name
                                or "fichier pcp" in col_name
                                or "header" in col_name
                                or "colonne source" in col_name
                            ):
                                ext_idx = idx
                        for row in df.values.tolist():
                            if not isinstance(row, list):
                                continue
                            code_value = row[code_idx] if code_idx < len(row) else None
                            name_value = row[name_idx] if name_idx < len(row) else None
                            ext_value = row[ext_idx] if ext_idx is not None and ext_idx < len(row) else None
                            _register(code_value, name_value, ext_value)
            finally:
                try:
                    xls.close()
                except Exception:
                    pass
        except Exception:
            pass

    if alias_map:
        return alias_map

    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = "Bassins" if "Bassins" in wb.sheetnames else ("Stations" if "Stations" in wb.sheetnames else None)
        if sheet_name:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header = list(next(rows_iter, []) or [])
            header_norm = [str(h or "").strip().lower() for h in header]
            code_idx = 0
            name_idx = 1 if len(header_norm) > 1 else 0
            ext_idx: Optional[int] = None
            for idx, col_name in enumerate(header_norm):
                if col_name in {"code", "id", "bassin_code", "station_code"}:
                    code_idx = idx
                if col_name in {"nom", "name", "bassin_nom", "station_nom", "nom station"}:
                    name_idx = idx
                if (
                    "nom station fichier" in col_name
                    or "station fichier" in col_name
                    or "fichier pcp" in col_name
                    or "header" in col_name
                    or "colonne source" in col_name
                ):
                    ext_idx = idx
            for row in rows_iter:
                row_list = list(row or [])
                code_value = row_list[code_idx] if code_idx < len(row_list) else None
                name_value = row_list[name_idx] if name_idx < len(row_list) else None
                ext_value = row_list[ext_idx] if ext_idx is not None and ext_idx < len(row_list) else None
                _register(code_value, name_value, ext_value)
        wb.close()
    except Exception:
        pass

    return alias_map


def _coerce_datetime_utc(value: Any) -> Optional[datetime]:
    return _coerce_datetime_utc_with_style(value, "AUTO")


MONTH_NAME_TO_NUMBER = {
    "jan": 1,
    "january": 1,
    "janv": 1,
    "janvier": 1,
    "feb": 2,
    "february": 2,
    "fev": 2,
    "fevr": 2,
    "fevrier": 2,
    "mar": 3,
    "march": 3,
    "mars": 3,
    "apr": 4,
    "april": 4,
    "avr": 4,
    "avril": 4,
    "may": 5,
    "mai": 5,
    "jun": 6,
    "june": 6,
    "juin": 6,
    "jul": 7,
    "july": 7,
    "juil": 7,
    "juillet": 7,
    "aug": 8,
    "august": 8,
    "aou": 8,
    "aout": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
    "decembre": 12,
}


def _parse_compact_month_name_datetime(text_value: str) -> Optional[datetime]:
    """
    Parse compact datetime strings like:
    - 01jan2026 0000
    - 1 jan 2026 00:00
    - 01-jan-2026 00:00:00
    """
    raw = str(text_value or "").strip().lower()
    if not raw:
        return None

    normalized = _fold_text(raw)
    normalized = re.sub(r"[._-]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    # compact token: 01jan2026 + optional time
    compact = re.match(r"^(\d{1,2})([a-z]+)(\d{4})(?:\s+(\d{2})(\d{2})(\d{2})?)?$", normalized)
    if compact:
        day = int(compact.group(1))
        month_name = compact.group(2)
        year = int(compact.group(3))
        month = MONTH_NAME_TO_NUMBER.get(month_name)
        if month is None:
            return None
        hour = int(compact.group(4) or 0)
        minute = int(compact.group(5) or 0)
        second = int(compact.group(6) or 0)
        try:
            return datetime(year, month, day, hour, minute, second)
        except ValueError:
            return None

    # spaced token: 01 jan 2026 + optional time
    spaced = re.match(
        r"^(\d{1,2})\s+([a-z]+)\s+(\d{4})(?:\s+(\d{1,2})(?::?(\d{2}))(?::?(\d{2}))?)?$",
        normalized,
    )
    if spaced:
        day = int(spaced.group(1))
        month_name = spaced.group(2)
        year = int(spaced.group(3))
        month = MONTH_NAME_TO_NUMBER.get(month_name)
        if month is None:
            return None
        hour = int(spaced.group(4) or 0)
        minute = int(spaced.group(5) or 0)
        second = int(spaced.group(6) or 0)
        try:
            return datetime(year, month, day, hour, minute, second)
        except ValueError:
            return None

    return None


def _detect_datetime_style(records: List[Dict[str, Any]], ts_col: str) -> Tuple[str, str]:
    """
    Detect dominant datetime format to avoid DD/MM vs MM/DD traps.
    Returns (style, confidence) where style in:
      ISO, YMD, DMY, MDY, AUTO
    """
    iso_like = 0
    ymd_like = 0
    dmy_like = 0
    mdy_like = 0
    month_name_like = 0
    dmy_strong = 0  # first token > 12
    mdy_strong = 0  # second token > 12
    sampled = 0

    for row in records:
        raw = row.get(ts_col)
        if raw is None:
            continue
        text_value = str(raw).strip()
        if not text_value:
            continue
        sampled += 1
        # Analyze a wider sample to avoid head-only ambiguity (e.g. up to 12/03).
        if sampled > 20000:
            break

        # ISO / YMD-like
        if re.match(r"^\d{4}-\d{2}-\d{2}", text_value):
            iso_like += 1
            ymd_like += 1
            continue
        if re.match(r"^\d{4}/\d{2}/\d{2}", text_value):
            ymd_like += 1
            continue

        if _parse_compact_month_name_datetime(text_value) is not None:
            month_name_like += 1
            continue

        slash_match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", text_value)
        if slash_match:
            a = int(slash_match.group(1))
            b = int(slash_match.group(2))
            if a > 12 and b <= 12:
                dmy_strong += 1
            elif b > 12 and a <= 12:
                mdy_strong += 1
            # fallback weak score by parse success
            try:
                datetime.strptime(text_value, "%d/%m/%Y %H:%M:%S")
                dmy_like += 1
            except ValueError:
                try:
                    datetime.strptime(text_value, "%d/%m/%Y %H:%M")
                    dmy_like += 1
                except ValueError:
                    try:
                        datetime.strptime(text_value, "%d/%m/%Y")
                        dmy_like += 1
                    except ValueError:
                        pass
            try:
                datetime.strptime(text_value, "%m/%d/%Y %H:%M:%S")
                mdy_like += 1
            except ValueError:
                try:
                    datetime.strptime(text_value, "%m/%d/%Y %H:%M")
                    mdy_like += 1
                except ValueError:
                    try:
                        datetime.strptime(text_value, "%m/%d/%Y")
                        mdy_like += 1
                    except ValueError:
                        pass

    if sampled == 0:
        return ("AUTO", "none")

    if iso_like > 0 and dmy_like == 0 and mdy_like == 0:
        return ("ISO", "high")

    if ymd_like > 0 and dmy_like == 0 and mdy_like == 0:
        return ("YMD", "high")

    if month_name_like > 0 and iso_like == 0 and ymd_like == 0:
        return ("DMY", "high")

    if dmy_strong > 0 and mdy_strong == 0:
        return ("DMY", "high")

    # Only pick MDY when evidence is very clear and DMY has no strong evidence.
    if mdy_strong > 0 and dmy_strong == 0 and mdy_like > (dmy_like * 1.2):
        return ("MDY", "high")

    if dmy_like > mdy_like:
        return ("DMY", "medium")
    if mdy_like > dmy_like and mdy_like > (dmy_like * 1.2):
        return ("MDY", "medium")

    # Ambiguous slash dates: prefer DMY for project ingestion context (fr-FR datasets).
    return ("DMY", "low")


def _coerce_datetime_utc_with_style(value: Any, style: str = "AUTO") -> Optional[datetime]:
    if value is None:
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        text_value = str(value).strip()
        if not text_value:
            return None

        normalized = text_value.replace("Z", "+00:00")
        style_upper = (style or "AUTO").strip().upper()
        parsed = None

        # 1) ISO / native first
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            parsed = None

        # 2) style-aware explicit parsing
        if parsed is None:
            parsed = _parse_compact_month_name_datetime(text_value)

        if parsed is None:
            if style_upper == "DMY":
                fmts = ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y")
            elif style_upper == "MDY":
                fmts = ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y")
            elif style_upper == "YMD":
                fmts = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d")
            elif style_upper == "ISO":
                fmts = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d")
            else:
                # AUTO fallback: prefer DMY before MDY to align with fr-FR ingestion habits
                fmts = (
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%d %H:%M",
                    "%Y-%m-%d",
                    "%d/%m/%Y %H:%M:%S",
                    "%d/%m/%Y %H:%M",
                    "%d/%m/%Y",
                    "%m/%d/%Y %H:%M:%S",
                    "%m/%d/%Y %H:%M",
                    "%m/%d/%Y",
                )

            for fmt in fmts:
                try:
                    parsed = datetime.strptime(text_value, fmt)
                    break
                except ValueError:
                    continue

        if parsed is None:
            return None
        dt = parsed

    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _swap_day_month(dt: datetime) -> Optional[datetime]:
    if dt.day > 12 or dt.month > 12 or dt.day == dt.month:
        return None
    try:
        return dt.replace(day=dt.month, month=dt.day)
    except ValueError:
        return None


def _sanitize_timestamp_sequence(
    parsed_records: List[Dict[str, Any]],
    ts_col: str,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """
    Timeline sanity correction:
    - Detect large jumps inconsistent with dominant cadence.
    - Try day/month swap on offending timestamp (typical MM/DD vs DD/MM corruption).
    - Drop unrecoverable outliers.
    """
    if len(parsed_records) < 3:
        return parsed_records, {"corrected_day_month_swap": 0, "dropped_outlier_timestamps": 0}

    corrected = 0
    dropped = 0
    output: List[Dict[str, Any]] = []
    positive_deltas: List[float] = []
    prev_dt: Optional[datetime] = None

    for row in parsed_records:
        raw_dt = row.get(ts_col)
        if not isinstance(raw_dt, datetime):
            dropped += 1
            continue

        candidate = raw_dt
        if prev_dt is not None:
            raw_delta = (candidate - prev_dt).total_seconds()
            for d in positive_deltas[-200:]:
                pass
            expected_step = median(positive_deltas[-200:]) if positive_deltas else 3600.0
            jump_threshold = max(expected_step * 8.0, 72.0 * 3600.0)

            if abs(raw_delta) > jump_threshold:
                swapped = _swap_day_month(candidate)
                if swapped is not None:
                    swapped_delta = (swapped - prev_dt).total_seconds()
                    if abs(swapped_delta) <= jump_threshold and abs(swapped_delta) < abs(raw_delta):
                        candidate = swapped
                        corrected += 1
                    else:
                        dropped += 1
                        continue
                else:
                    dropped += 1
                    continue

        row_copy = dict(row)
        row_copy[ts_col] = candidate
        output.append(row_copy)

        if prev_dt is not None:
            d = (candidate - prev_dt).total_seconds()
            if d > 0 and d <= 15 * 24 * 3600:
                positive_deltas.append(d)
        prev_dt = candidate

    return output, {"corrected_day_month_swap": corrected, "dropped_outlier_timestamps": dropped}


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text_value = str(value).strip()
    if not text_value:
        return None

    try:
        return float(text_value.replace(",", "."))
    except ValueError:
        return None


def _fold_text(value: Any) -> str:
    raw = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", raw)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_identifier(value: Any) -> str:
    folded = _fold_text(value)
    cleaned = re.sub(r"[^a-z0-9]+", "_", folded)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned


def _identifier_tokens(value: Any) -> List[str]:
    normalized = _normalize_identifier(value)
    if not normalized:
        return []
    return [token for token in normalized.split("_") if token]


def _ordered_unique(values: List[str]) -> List[str]:
    seen: Set[str] = set()
    output: List[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return output


def _infer_source_code_from_column(column_name: str) -> Optional[str]:
    normalized = _normalize_identifier(column_name)
    if not normalized:
        return None

    tokens = _identifier_tokens(normalized)
    for token in tokens:
        mapped = SOURCE_HINT_TOKENS.get(token)
        if mapped:
            return mapped

    if re.search(r"(^|_)(arome)($|_)", normalized):
        return "AROME"
    if re.search(r"(^|_)(ecmwf)($|_)", normalized):
        return "ECMWF"
    if re.search(r"(^|_)(obs|observe|observation|historique|history)($|_)", normalized):
        return "OBS"
    if re.search(r"(^|_)(sim|simu|simule|simulation|simulated)($|_)", normalized):
        return "SIM"
    return None


def _infer_source_code_from_text(value: Any, *, require_source_keyword: bool = False) -> Optional[str]:
    raw = str(value or "").strip()
    if not raw:
        return None

    folded = _fold_text(raw)
    if require_source_keyword and "source" not in folded:
        return None

    code_match = re.search(r"\b(OBS|SIM|AROME|ECMWF)\b", raw, flags=re.IGNORECASE)
    if code_match:
        return code_match.group(1).upper()

    return _infer_source_code_from_column(raw)


def _extract_source_codes_from_text(value: Any, *, require_source_keyword: bool = False) -> List[str]:
    raw = str(value or "").strip()
    if not raw:
        return []

    folded = _fold_text(raw)
    if require_source_keyword and "source" not in folded:
        return []

    detected: List[str] = []
    for match in re.finditer(r"\b(OBS|SIM|AROME|ECMWF)\b", raw, flags=re.IGNORECASE):
        detected.append(match.group(1).upper())

    if detected:
        return _ordered_unique(detected)

    inferred = _infer_source_code_from_column(raw)
    if inferred:
        return [inferred]
    return []


def _extract_source_codes_from_columns(columns: List[str]) -> List[str]:
    detected: List[str] = []
    for column_name in columns:
        source_hint = _infer_source_code_from_column(column_name or "")
        if source_hint:
            detected.append(source_hint)
    return _ordered_unique(detected)


def _extract_source_codes_from_file(file_path: str, filename: str) -> List[str]:
    probe_cells: List[str] = []
    lower = (filename or "").lower()

    try:
        if lower.endswith(".csv"):
            with open(file_path, "r", encoding="utf-8-sig", errors="replace") as handle:
                for _ in range(20):
                    line = handle.readline()
                    if not line:
                        break
                    probe_cells.append(line.strip())
        else:
            pd = _get_pandas()
            if pd is not None:
                try:
                    preview = pd.read_excel(file_path, header=None, nrows=10)
                    for row in preview.values.tolist():
                        for value in row[:12]:
                            if not _is_empty_cell(value):
                                probe_cells.append(str(value))
                except Exception:
                    pass

            if not probe_cells:
                load_workbook = None
                try:
                    from openpyxl import load_workbook as _load_wb  # type: ignore
                    load_workbook = _load_wb
                except Exception:
                    fallback_paths = []
                    conda_prefix = os.environ.get("CONDA_PREFIX")
                    if conda_prefix:
                        fallback_paths.append(os.path.join(conda_prefix, "Lib", "site-packages"))
                    if os.name == "nt":
                        fallback_paths.append(r"C:\anaconda\Lib\site-packages")
                        fallback_paths.append(r"C:\ProgramData\anaconda3\Lib\site-packages")

                    for path in fallback_paths:
                        if path and os.path.isdir(path) and path not in sys.path:
                            sys.path.append(path)
                        try:
                            from openpyxl import load_workbook as _load_wb  # type: ignore
                            load_workbook = _load_wb
                            break
                        except Exception:
                            continue

                if load_workbook is not None:
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if ridx > 10:
                            break
                        for value in list(row)[:12]:
                            if not _is_empty_cell(value):
                                probe_cells.append(str(value))
                    wb.close()
    except Exception:
        return []

    detected: List[str] = []
    for cell in probe_cells:
        detected.extend(_extract_source_codes_from_text(cell, require_source_keyword=True))

    if detected:
        return _ordered_unique([code for code in detected if code in KNOWN_SOURCE_CODES])

    for cell in probe_cells:
        detected.extend(_extract_source_codes_from_text(cell, require_source_keyword=False))

    return _ordered_unique([code for code in detected if code in KNOWN_SOURCE_CODES])


def _extract_source_code_from_file(file_path: str, filename: str) -> Optional[str]:
    codes = _extract_source_codes_from_file(file_path, filename)
    return codes[0] if codes else None


def _strip_source_tokens(value: Any) -> str:
    tokens = [
        token for token in _identifier_tokens(value)
        if token not in SOURCE_HINT_TOKENS and token not in SOURCE_DECORATION_TOKENS
    ]
    return "_".join(tokens)


def _station_name_stem(value: Any) -> str:
    normalized = _normalize_identifier(value)
    if normalized:
        normalized = normalized.replace("d_bit", "debit")
        normalized = normalized.replace("de_bit", "debit")
    tokens = [
        token for token in _identifier_tokens(normalized or value)
        if token not in SOURCE_HINT_TOKENS
        and token not in SOURCE_DECORATION_TOKENS
        and token not in STATION_DECORATION_TOKENS
    ]
    stem = "_".join(tokens).strip("_")
    return re.sub(r"_+", "_", stem)


def _alias_lookup_keys(value: Any) -> List[str]:
    base = _entity_lookup_keys(value)
    stem = _station_name_stem(value)
    if stem:
        base.append(stem)
    return _ordered_unique(base)


def _infer_variable_family(value: Any) -> Optional[str]:
    normalized = _normalize_identifier(value)
    if not normalized:
        return None
    tokens = set(_identifier_tokens(normalized))
    if tokens.intersection(PRECIP_HINT_TOKENS):
        return "precip"
    if tokens.intersection(HYDRO_HINT_TOKENS):
        return "hydro"

    if any(key in normalized for key in ("precip", "pluie", "rain", "snow")):
        return "precip"
    if any(key in normalized for key in ("flow", "debit", "inflow", "apport", "volume", "cote", "lacher", "solide", "solid")):
        return "hydro"
    return None


def _suggest_source_codes(detected_families: Set[str], detected_sources: Set[str]) -> List[str]:
    ordered = ["OBS", "AROME", "ECMWF", "SIM"]
    if not detected_families:
        if detected_sources:
            return [code for code in ordered if code in detected_sources]
        return ["OBS"]

    if detected_families == {"precip"}:
        base = ["OBS", "AROME", "ECMWF"]
    elif detected_families == {"hydro"}:
        base = ["OBS", "SIM"]
    else:
        base = ["OBS", "AROME", "ECMWF", "SIM"]

    for code in ordered:
        if code in detected_sources and code not in base:
            base.append(code)
    return base


def _allowed_sources_for_family(family: Optional[str]) -> List[str]:
    normalized = (family or "").strip().lower()
    if normalized == "precip":
        return ["OBS", "AROME", "ECMWF"]
    if normalized == "hydro":
        return ["OBS", "SIM"]
    return ["OBS", "SIM", "AROME", "ECMWF"]


def _is_source_compatible_with_family(source_code: str, family: Optional[str]) -> bool:
    src = (source_code or "").strip().upper()
    if not src:
        return False
    return src in _allowed_sources_for_family(family)


def _column_variable_candidates(column_name: str) -> List[str]:
    if not column_name:
        return []

    candidates: List[str] = []
    without_parentheses = re.sub(r"\(.*?\)", "", column_name).strip()
    for raw_candidate in (column_name, without_parentheses):
        normalized = _normalize_identifier(raw_candidate)
        if not normalized:
            continue
        candidates.append(normalized)
        without_source = _strip_source_tokens(normalized)
        if without_source and without_source != normalized:
            candidates.append(without_source)

    for candidate in list(candidates):
        if candidate.endswith("_value"):
            candidates.append(candidate[:-6])
        if candidate.endswith("_valeur"):
            candidates.append(candidate[:-7])
        tokens = [t for t in candidate.split("_") if t]
        if tokens and tokens[-1] in {"value", "valeur", "val", "data", "donnee", "donnees"}:
            trimmed = "_".join(tokens[:-1])
            if trimmed:
                candidates.append(trimmed)

    return _ordered_unique(candidates)


def _expand_variable_aliases(candidates: List[str]) -> List[str]:
    ordered: List[str] = []
    queue = list(candidates)
    seen: Set[str] = set()

    while queue:
        candidate = queue.pop(0)
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        ordered.append(candidate)

        alias = VARIABLE_COLUMN_ALIASES.get(candidate)
        if alias and alias not in seen:
            queue.append(alias)

    return ordered


def _entity_lookup_keys(value: Any) -> List[str]:
    raw = str(value or "").strip()
    if not raw:
        return []

    keys: List[str] = [raw.lower()]
    normalized = _normalize_identifier(raw)
    if normalized:
        keys.append(normalized)
        stripped = _strip_source_tokens(normalized)
        if stripped and stripped != normalized:
            keys.append(stripped)

        # DGM template compatibility:
        # "DGM-1_arome" -> "dgm_1" -> add "1" to match numeric basin codes.
        dgm_match = re.match(r"^dgm[_-]?0*([0-9]+)$", stripped or normalized)
        if dgm_match:
            num = dgm_match.group(1)
            keys.append(num)
            keys.append(f"dgm_{num}")
            keys.append(f"dgm-{num}")

    numeric_clean = raw.lower()
    if numeric_clean.endswith(".0") and numeric_clean.replace(".", "", 1).isdigit():
        keys.append(numeric_clean[:-2])
    return _ordered_unique(keys)


async def _load_variable_lookup(db: AsyncSession) -> Dict[str, Dict[str, Any]]:
    v_res = await db.execute(text("SELECT variable_id, code, label, unit FROM ref.variable"))
    lookup: Dict[str, Dict[str, Any]] = {}
    for row in v_res.mappings().all():
        payload = {
            "variable_id": row["variable_id"],
            "variable_code": str(row["code"]),
            "variable_label": str(row.get("label") or ""),
            "unit": str(row.get("unit") or ""),
        }
        candidates = _column_variable_candidates(payload["variable_code"]) + _column_variable_candidates(payload["variable_label"])
        for candidate in _expand_variable_aliases(candidates):
            lookup.setdefault(candidate, payload)
    return lookup


def _resolve_variable_from_column(column_name: str, variable_lookup: Dict[str, Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    candidates = _expand_variable_aliases(_column_variable_candidates(column_name))
    for candidate in candidates:
        payload = variable_lookup.get(candidate)
        if payload is not None:
            return payload
    return None


def _is_quality_flag_column(column_name: str) -> bool:
    normalized = _normalize_identifier(column_name)
    if not normalized:
        return False
    if normalized in QUALITY_FLAG_COLUMN_TOKENS:
        return True
    tokens = set(_identifier_tokens(normalized))
    if {"quality", "flag"}.issubset(tokens):
        return True
    if {"qualite", "flag"}.issubset(tokens):
        return True
    return normalized.startswith("qc_")


def _pick_simple_value_column(data_columns: List[str], variable_code: Optional[str]) -> Optional[str]:
    if not data_columns:
        return None

    normalized_by_col = {col: _normalize_identifier(col) for col in data_columns}

    preferred_generic = {"value", "valeur", "val", "data", "donnee", "donnees"}
    for col in data_columns:
        normalized = normalized_by_col.get(col, "")
        if normalized in preferred_generic and not _is_quality_flag_column(col):
            return col

    if variable_code:
        variable_candidates = set(_expand_variable_aliases(_column_variable_candidates(variable_code)))
        for col in data_columns:
            normalized = normalized_by_col.get(col, "")
            if normalized in variable_candidates and not _is_quality_flag_column(col):
                return col

    for col in data_columns:
        if not _is_quality_flag_column(col):
            return col

    return data_columns[0]


async def _load_source_ids_map(db: AsyncSession) -> Dict[str, Any]:
    src_res = await db.execute(text("SELECT source_id, code FROM ref.source"))
    source_map: Dict[str, Any] = {}
    for row in src_res.mappings().all():
        code = str(row["code"]).upper()
        source_map[code] = row["source_id"]
    return source_map


async def _load_entity_map(db: AsyncSession, entity_type: str) -> Dict[str, Dict[str, str]]:
    entity_map: Dict[str, Dict[str, str]] = {}

    if entity_type == "bassins":
        # Primary lookup: direct basin code/name.
        basin_res = await db.execute(
            text(
                """
                SELECT
                    b.basin_id::text AS entity_id,
                    b.code::text AS code,
                    b.name AS name
                FROM geo.basin b
                """
            )
        )
        for row in basin_res.mappings().all():
            payload = {
                "id": str(row["entity_id"]),
                "type": "Bassin",
                "code": str(row.get("code") or ""),
                "name": str(row.get("name") or ""),
            }
            for key in _entity_lookup_keys(row.get("code")):
                entity_map.setdefault(key, payload)
            for key in _entity_lookup_keys(row.get("name")):
                entity_map.setdefault(key, payload)

        # Fallback lookup for DGM templates that may contain barrage labels:
        # map barrage station code/name to its basin_id.
        station_res = await db.execute(
            text(
                """
                SELECT
                    s.basin_id::text AS entity_id,
                    s.code::text AS code,
                    s.name AS name
                FROM geo.station s
                WHERE s.basin_id IS NOT NULL
                """
            )
        )
        for row in station_res.mappings().all():
            payload = {
                "id": str(row["entity_id"]),
                "type": "Bassin",
                "code": str(row.get("code") or ""),
                "name": str(row.get("name") or ""),
            }
            for key in _entity_lookup_keys(row.get("code")):
                entity_map.setdefault(key, payload)
            for key in _entity_lookup_keys(row.get("name")):
                entity_map.setdefault(key, payload)
        return entity_map

    station_res = await db.execute(
        text(
            """
            SELECT
                s.station_id::text AS entity_id,
                s.code::text AS code,
                s.name AS name,
                COALESCE(NULLIF(TRIM(s.station_type), ''), 'Station') AS station_type
            FROM geo.station s
            """
        )
    )
    for row in station_res.mappings().all():
        payload = {
            "id": str(row["entity_id"]),
            "type": str(row.get("station_type") or "Station"),
            "code": str(row.get("code") or ""),
            "name": str(row.get("name") or ""),
        }
        for key in _entity_lookup_keys(row.get("code")):
            entity_map.setdefault(key, payload)
        for key in _entity_lookup_keys(row.get("name")):
            entity_map.setdefault(key, payload)

    return entity_map


async def _load_entity_map_safe(db: AsyncSession, entity_type: str) -> Dict[str, Dict[str, str]]:
    try:
        return await _load_entity_map(db, entity_type)
    except Exception as exc:
        print(f"Warning: _load_entity_map failed ({exc}), falling back to legacy entity lookup.")
        entity_map: Dict[str, Dict[str, str]] = {}
        if entity_type == "bassins":
            basin_queries = [
                "SELECT basin_id as id, code, name FROM geo.basin",
                "SELECT basin_id as id, code, nom as name FROM geo.basin",
                "SELECT basin_id as id, basin_code as code, name FROM geo.basin",
                "SELECT basin_id as id, basin_code as code, nom as name FROM geo.basin",
                "SELECT basin_id as id, NULL::text as code, name FROM geo.basin",
                "SELECT basin_id as id, NULL::text as code, nom as name FROM geo.basin",
            ]
            rows = []
            for sql in basin_queries:
                try:
                    res = await db.execute(text(sql))
                    rows = res.mappings().all()
                    if rows is not None:
                        break
                except Exception:
                    continue
            for row in rows or []:
                payload = {
                    "id": str(row.get("id")),
                    "type": "Bassin",
                    "code": str(row.get("code") or ""),
                    "name": str(row.get("name") or ""),
                }
                for key in _entity_lookup_keys(row.get("code")):
                    entity_map.setdefault(key, payload)
                for key in _entity_lookup_keys(row.get("name")):
                    entity_map.setdefault(key, payload)
        else:
            station_queries = [
                "SELECT station_id as id, code, name, station_type as type FROM geo.station",
                "SELECT station_id as id, code, nom as name, station_type as type FROM geo.station",
                "SELECT station_id as id, station_code as code, name, station_type as type FROM geo.station",
                "SELECT station_id as id, station_code as code, nom as name, station_type as type FROM geo.station",
                "SELECT station_id as id, NULL::text as code, name, station_type as type FROM geo.station",
                "SELECT station_id as id, NULL::text as code, nom as name, station_type as type FROM geo.station",
            ]
            rows = []
            for sql in station_queries:
                try:
                    res = await db.execute(text(sql))
                    rows = res.mappings().all()
                    if rows is not None:
                        break
                except Exception:
                    continue
            for row in rows or []:
                payload = {
                    "id": str(row.get("id")),
                    "type": str(row.get("type") or "Station"),
                    "code": str(row.get("code") or ""),
                    "name": str(row.get("name") or ""),
                }
                for key in _entity_lookup_keys(row.get("code")):
                    entity_map.setdefault(key, payload)
                for key in _entity_lookup_keys(row.get("name")):
                    entity_map.setdefault(key, payload)
        return entity_map


def _entity_field(entity: Any, field: str, default: Optional[str] = None) -> Optional[str]:
    if entity is None:
        return default
    if isinstance(entity, dict):
        value = entity.get(field, default)
        return default if value is None else str(value)
    value = getattr(entity, field, default)
    return default if value is None else str(value)


def _safe_remove_tmp(path: str, retries: int = 5, delay: float = 0.15) -> None:
    if not path:
        return
    for attempt in range(retries):
        try:
            if os.path.exists(path):
                os.remove(path)
            return
        except PermissionError:
            gc.collect()
            if attempt == retries - 1:
                print(f"Warning: unable to remove temp file (locked): {path}")
                return
            time.sleep(delay * (attempt + 1))
        except FileNotFoundError:
            return
        except Exception as exc:
            print(f"Warning: temp cleanup failed for {path}: {exc}")
            return


async def _ensure_import_run_id(
    db: AsyncSession,
    source_id: Any,
    source_code: str,
    run_cache: Dict[str, Any],
) -> Any:
    source_key = str(source_code).upper()
    cached = run_cache.get(source_key)
    if cached:
        return cached

    run_label = f"Import_{source_key}_{datetime.utcnow().strftime('%Y%m%d')}"
    run_res = await db.execute(text("SELECT run_id FROM ref.run WHERE label = :label"), {"label": run_label})
    run_row = run_res.first()
    if run_row:
        run_id = run_row[0]
    else:
        run_id_res = await db.execute(
            text(
                """
                INSERT INTO ref.run (label, source_id, run_time)
                VALUES (:label, :source_id, :run_time)
                RETURNING run_id
                """
            ),
            {"label": run_label, "source_id": source_id, "run_time": datetime.utcnow()},
        )
        run_id = run_id_res.first()[0]

    run_cache[source_key] = run_id
    return run_id


async def _resolve_source_payload(
    db: AsyncSession,
    requested_source_code: str,
    source_ids_map: Dict[str, Any],
    run_cache: Dict[str, Any],
    *,
    column_name: Optional[str] = None,
    variable_hint: Optional[str] = None,
) -> Tuple[str, Any, Any]:
    requested = (requested_source_code or "OBS").strip().upper()
    if requested != "AUTO":
        if requested not in source_ids_map:
            raise HTTPException(404, f"Source {requested} not found")
        source_id = source_ids_map[requested]
        run_id = await _ensure_import_run_id(db, source_id, requested, run_cache)
        return requested, source_id, run_id

    family = _infer_variable_family(variable_hint or column_name or "")
    inferred = _infer_source_code_from_column(column_name or "")
    if not inferred:
        if family == "hydro":
            inferred = "OBS"
        elif family == "precip":
            inferred = "OBS"
        else:
            inferred = "OBS"
    elif not _is_source_compatible_with_family(inferred, family):
        inferred = "OBS"

    if inferred not in source_ids_map:
        if family == "precip":
            fallback_order = ["OBS", "AROME", "ECMWF", "SIM"]
        elif family == "hydro":
            fallback_order = ["OBS", "SIM", "AROME", "ECMWF"]
        else:
            fallback_order = ["OBS", "SIM", "AROME", "ECMWF"]
        fallback = next((code for code in fallback_order if code in source_ids_map), None)
        if not fallback:
            raise HTTPException(404, "No compatible source found in ref.source")
        inferred = fallback

    source_id = source_ids_map[inferred]
    run_id = await _ensure_import_run_id(db, source_id, inferred, run_cache)
    return inferred, source_id, run_id


@router.get("/timeseries/sources")
async def list_sources(
    db: AsyncSession = Depends(get_db)
):
    """Get list of available data sources"""
    try:
        query = text("SELECT code, label FROM ref.source ORDER BY source_id")
        result = await db.execute(query)
        rows = result.mappings().all()
        return [dict(row) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching sources: {str(e)}")

@router.post("/timeseries/analyze")
async def analyze_timeseries_file(
    file: UploadFile = File(...),
    entity_type: str = Form("stations"),
    import_mode: str = Form("multi_station"),
    station_id: Optional[str] = Form(None),
    variable_code: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    """
    Analyze file content before import.
    Returns details including data preview.
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file.filename.split('.')[-1]}") as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
            
        try:
            columns, records = _load_tabular_data(tmp_path, file.filename)
            columns = [str(c).strip() for c in columns]
            if not columns:
                return {"status": "error", "message": "File is empty or unreadable", "columns": []}
            file_source_hints = _extract_source_codes_from_file(tmp_path, file.filename)
            file_source_hint = file_source_hints[0] if file_source_hints else None
            
            # Detect timestamp
            ts_col = next((c for c in columns if c.lower() in TIMESTAMP_KEYWORDS), None)
            
            entity_map = await _load_entity_map_safe(db, entity_type)
            try:
                template_alias_map = _load_template_entity_alias_map(tmp_path, file.filename)
            except Exception as alias_exc:
                print(f"Warning: template alias map failed ({alias_exc})")
                template_alias_map = {}
            variable_lookup = await _load_variable_lookup(db)

            found_entities = []
            unknown_columns = []
            unmatched_entities: List[str] = []

            # Statistics
            rows_count = len(records)
            start_date = None
            end_date = None
            preview_data = []

            if ts_col:
                data_columns = [
                    col for col in columns
                    if col != ts_col and not col.lower().startswith("unnamed")
                ]
                detected_dt_style, detected_dt_confidence = _detect_datetime_style(records, ts_col)
                detected_source_set: Set[str] = set()
                if file_source_hints:
                    detected_source_set.update(file_source_hints)
                detected_family_set: Set[str] = set()
                column_analysis: List[Dict[str, Any]] = []

                parsed_records_for_timeline: List[Dict[str, Any]] = []
                for row in records:
                    ts_val = _coerce_datetime_utc_with_style(row.get(ts_col), detected_dt_style)
                    if ts_val is None:
                        continue
                    copied = dict(row)
                    copied[ts_col] = ts_val
                    parsed_records_for_timeline.append(copied)

                sanitized_records, ts_sanity = _sanitize_timestamp_sequence(parsed_records_for_timeline, ts_col)
                valid_times = [r[ts_col] for r in sanitized_records if isinstance(r.get(ts_col), datetime)]
                if valid_times:
                    start_date = min(valid_times).isoformat()
                    end_date = max(valid_times).isoformat()

                # Generate preview (first 5 rows)
                for row in records[:5]:
                    preview_row = {}
                    for col in columns:
                        val = row.get(col)
                        preview_row[col] = "" if _is_empty_cell(val) else str(val)
                    preview_data.append(preview_row)

                normalized_mode = (import_mode or "multi_station").strip().lower()
                if normalized_mode not in {"simple", "multi_station", "multi_variable"}:
                    normalized_mode = "multi_station"

                if normalized_mode == "simple":
                    selected_target = None
                    if station_id:
                        try:
                            if entity_type == "bassins":
                                target_res = await db.execute(
                                    text(
                                        """
                                        SELECT
                                            b.basin_id::text AS entity_id,
                                            COALESCE(NULLIF(TRIM(b.name), ''), b.basin_id::text) AS name,
                                            NULLIF(TRIM(b.code::text), '') AS code,
                                            'Bassin'::text AS station_type
                                        FROM geo.basin b
                                        WHERE b.basin_id = CAST(:id AS UUID)
                                        """
                                    ),
                                    {"id": station_id},
                                )
                            else:
                                target_res = await db.execute(
                                    text(
                                        """
                                        SELECT
                                            s.station_id::text AS entity_id,
                                            COALESCE(NULLIF(TRIM(s.name), ''), s.station_id::text) AS name,
                                            NULLIF(TRIM(s.code::text), '') AS code,
                                            COALESCE(NULLIF(TRIM(s.station_type), ''), 'Station') AS station_type
                                        FROM geo.station s
                                        WHERE s.station_id = CAST(:id AS UUID)
                                        """
                                    ),
                                    {"id": station_id},
                                )
                            target_row = target_res.mappings().first()
                            if target_row:
                                selected_target = {
                                    "column": "(cible)",
                                    "matched_station": target_row["name"],
                                    "station_code": target_row["code"],
                                    "type": target_row["station_type"],
                                }
                        except Exception:
                            selected_target = None

                    selected_variable_payload = None
                    if variable_code:
                        selected_variable_payload = _resolve_variable_from_column(variable_code, variable_lookup)
                        if selected_variable_payload is None:
                            selected_variable_payload = {
                                "variable_code": variable_code,
                                "variable_label": variable_code,
                                "unit": None,
                            }

                    value_col = _pick_simple_value_column(data_columns, variable_code)
                    quality_columns = {
                        col for col in data_columns
                        if col != value_col and _is_quality_flag_column(col)
                    }

                    for col in data_columns:
                        if col == value_col:
                            variable_code_hint = (
                                selected_variable_payload.get("variable_code")
                                if selected_variable_payload
                                else (variable_code or col)
                            )
                            family_hint = _infer_variable_family(variable_code_hint)
                            source_hint = _infer_source_code_from_column(col)
                            if source_hint and not _is_source_compatible_with_family(source_hint, family_hint):
                                source_hint = "OBS"
                            if source_hint:
                                detected_source_set.add(source_hint)
                            if family_hint:
                                detected_family_set.add(family_hint)
                            column_analysis.append(
                                {
                                    "column": col,
                                    "kind": "value",
                                    "family_hint": family_hint,
                                    "source_hint": source_hint,
                                    "variable_code": selected_variable_payload.get("variable_code") if selected_variable_payload else variable_code,
                                    "variable_label": selected_variable_payload.get("variable_label") if selected_variable_payload else variable_code,
                                    "unit": selected_variable_payload.get("unit") if selected_variable_payload else None,
                                    "matched_station": selected_target["matched_station"] if selected_target else None,
                                    "matched_station_code": selected_target["station_code"] if selected_target else None,
                                }
                            )
                            continue

                        if col in quality_columns:
                            column_analysis.append(
                                {
                                    "column": col,
                                    "kind": "quality_flag",
                                    "family_hint": None,
                                    "source_hint": None,
                                    "variable_code": None,
                                    "variable_label": None,
                                    "unit": None,
                                    "matched_station": selected_target["matched_station"] if selected_target else None,
                                    "matched_station_code": selected_target["station_code"] if selected_target else None,
                                }
                            )
                            continue

                        variable_payload = _resolve_variable_from_column(col, variable_lookup)
                        variable_code_hint = variable_payload.get("variable_code") if variable_payload else None
                        family_hint = _infer_variable_family(variable_code_hint or col)
                        source_hint = _infer_source_code_from_column(col)
                        if source_hint and not _is_source_compatible_with_family(source_hint, family_hint):
                            source_hint = "OBS"
                        if source_hint:
                            detected_source_set.add(source_hint)
                        if family_hint:
                            detected_family_set.add(family_hint)

                        if variable_payload is None:
                            unknown_columns.append(col)

                        column_analysis.append(
                            {
                                "column": col,
                                "kind": "variable" if variable_payload else "unknown",
                                "family_hint": family_hint,
                                "source_hint": source_hint,
                                "variable_code": variable_payload.get("variable_code") if variable_payload else None,
                                "variable_label": variable_payload.get("variable_label") if variable_payload else None,
                                "unit": variable_payload.get("unit") if variable_payload else None,
                                "matched_station": selected_target["matched_station"] if selected_target else None,
                                "matched_station_code": selected_target["station_code"] if selected_target else None,
                            }
                        )

                    found_entities = [selected_target] if selected_target else []

                else:
                    for col in data_columns:
                        variable_payload = _resolve_variable_from_column(col, variable_lookup)
                        variable_code_hint = variable_payload.get("variable_code") if variable_payload else None
                        family_hint = _infer_variable_family(variable_code_hint or col)
                        source_hint = _infer_source_code_from_column(col)
                        if source_hint and not _is_source_compatible_with_family(source_hint, family_hint):
                            source_hint = "OBS"

                        if source_hint:
                            detected_source_set.add(source_hint)
                        if family_hint:
                            detected_family_set.add(family_hint)

                        matched_entity = None
                        for key in _alias_lookup_keys(col):
                            if key in entity_map:
                                matched_entity = entity_map[key]
                                break
                        if matched_entity is None and template_alias_map:
                            alias_name = None
                            for key in _alias_lookup_keys(col):
                                alias_name = template_alias_map.get(key)
                                if alias_name:
                                    break
                            if alias_name:
                                for alias_key in _entity_lookup_keys(alias_name):
                                    if alias_key in entity_map:
                                        matched_entity = entity_map[alias_key]
                                        break

                        if matched_entity is not None:
                            ent = matched_entity
                            found_entities.append({
                                "column": col,
                                "matched_station": _entity_field(ent, "name", col) or col,
                                "station_code": _entity_field(ent, "code"),
                                "type": _entity_field(ent, "type", "Station") or "Station"
                            })
                        elif variable_payload is None:
                            unknown_columns.append(col)
                            unmatched_entities.append(col)

                        kind = "station_or_bassin" if matched_entity is not None else ("variable" if variable_payload else "unknown")
                        column_analysis.append(
                            {
                                "column": col,
                                "kind": kind,
                                "family_hint": family_hint,
                                "source_hint": source_hint,
                                "variable_code": variable_payload.get("variable_code") if variable_payload else None,
                                "variable_label": variable_payload.get("variable_label") if variable_payload else None,
                                "unit": variable_payload.get("unit") if variable_payload else None,
                                "matched_station": (_entity_field(matched_entity, "name", col) or col) if matched_entity is not None else None,
                                "matched_station_code": _entity_field(matched_entity, "code") if matched_entity is not None else None,
                            }
                        )

                return {
                    "status": "success",
                    "filename": file.filename,
                    "rows_count": rows_count,
                    "time_column": ts_col,
                    "start_date": start_date,
                    "end_date": end_date,
                    "stations_found": len(found_entities),
                    "stations_details": found_entities,
                    "unknown_columns": unknown_columns,
                    "unmatched_entities": unmatched_entities,
                    "preview": preview_data,
                    "columns": columns,
                    "file_source_hint": file_source_hint,
                    "file_source_hints": file_source_hints,
                    "detected_sources": [c for c in ["OBS", "AROME", "ECMWF", "SIM"] if c in detected_source_set],
                    "detected_families": [f for f in ["precip", "hydro"] if f in detected_family_set],
                    "detected_datetime_format": detected_dt_style,
                    "detected_datetime_confidence": detected_dt_confidence,
                    "timestamp_sanity": ts_sanity,
                    "datetime_format_warning": (
                        "Format date ambigu detecte (ex: 03/04/2026). Le systeme applique DMY (jour/mois) par defaut."
                        if detected_dt_confidence == "low"
                        else None
                    ),
                    "suggested_sources": _suggest_source_codes(detected_family_set, detected_source_set),
                    "auto_source_supported": any(item.get("source_hint") for item in column_analysis),
                    "column_analysis": column_analysis,
                    "source_policy": {
                        "AUTO": "Scan des colonnes (ex: variable_obs/sim/arome/ecmwf_unite)",
                        "OBS": "Appliquer OBS a toutes les variables detectees",
                        "SIM": "Appliquer SIM aux variables hydrologiques (debit/apport/volume), jamais a la pluie",
                        "AROME": "Appliquer AROME uniquement a la pluie",
                        "ECMWF": "Appliquer ECMWF uniquement a la pluie",
                    },
                }
            else:
                 return {
                    "status": "error",
                    "message": "Timestamp column not found"
                }

        finally:
            _safe_remove_tmp(tmp_path)
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Analysis failed: {str(e)}")

@router.post("/timeseries/upload")
async def upload_timeseries(
    file: UploadFile = File(...),
    import_mode: str = Form("simple"),
    replace_existing: str = Form("false"),
    station_id: Optional[str] = Form(None),
    variable_code: Optional[str] = Form(None),
    source_code: str = Form("AUTO"),
    entity_type: str = Form("stations"),
    db: AsyncSession = Depends(get_db)
):
    """
    Upload time series data.
    - source_code: OBS, SIM, AROME, ECMWF, or AUTO
    - La source est prioritairement lue depuis le fichier (metadonnees/canevas),
      puis depuis le parametre `source_code`, puis inferée depuis les colonnes.
    """
    try:
        # Save temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file.filename.split('.')[-1]}") as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        
        try:
            # Parse file (uses pandas when available, else builtin csv/openpyxl fallback)
            columns, records = _load_tabular_data(tmp_path, file.filename)
            columns = [str(c).strip() for c in columns]
            if not columns:
                raise HTTPException(400, "File is empty or unreadable")
            
            # Detect timestamp column
            ts_col = next((c for c in columns if c.lower() in TIMESTAMP_KEYWORDS), None)
            if not ts_col:
                # Try first column
                ts_col = columns[0]

            detected_dt_style, detected_dt_confidence = _detect_datetime_style(records, ts_col)
            
            # Parse timestamps and keep only valid rows
            parsed_records = []
            for row in records:
                ts_val = _coerce_datetime_utc_with_style(row.get(ts_col), detected_dt_style)
                if ts_val is None:
                    continue
                copied = dict(row)
                copied[ts_col] = ts_val
                parsed_records.append(copied)

            parsed_records, ts_sanity = _sanitize_timestamp_sequence(parsed_records, ts_col)
             
            if not parsed_records:
                 raise HTTPException(400, "File is empty or no valid timestamps found")

            file_source_hints = _extract_source_codes_from_file(tmp_path, file.filename)
            file_source_hint = file_source_hints[0] if file_source_hints else None
            column_source_hints = _extract_source_codes_from_columns(columns)

            requested_source_input = (source_code or "").strip().upper()
            if len(file_source_hints) > 1:
                requested_source_code = "AUTO"
                source_resolution_mode = "file_metadata_multi_source"
            elif file_source_hint:
                requested_source_code = file_source_hint
                source_resolution_mode = "file_metadata"
            elif requested_source_input and requested_source_input != "AUTO":
                requested_source_code = requested_source_input
                source_resolution_mode = "request_param"
            elif len(column_source_hints) == 1:
                requested_source_code = column_source_hints[0]
                source_resolution_mode = "column_hint"
            else:
                requested_source_code = "AUTO"
                source_resolution_mode = "auto_by_column"

            source_ids_map = await _load_source_ids_map(db)
            if requested_source_code != "AUTO" and requested_source_code not in source_ids_map:
                raise HTTPException(404, f"Source {requested_source_code} not found")
            run_cache: Dict[str, Any] = {}
            policy_skipped_columns: List[Dict[str, Any]] = []

            records_count = 0
            
            if import_mode == "simple":
                if not station_id or not variable_code:
                    raise HTTPException(status_code=400, detail="Station and Variable required for simple mode")

                data_columns = [
                    c for c in columns
                    if c != ts_col and not c.lower().startswith("unnamed")
                ]
                if not data_columns:
                    raise HTTPException(status_code=400, detail="No value column found")

                quality_columns = [c for c in data_columns if _is_quality_flag_column(c)]
                value_columns = [c for c in data_columns if c not in quality_columns]
                if not value_columns:
                    fallback_value_col = _pick_simple_value_column(data_columns, variable_code)
                    if fallback_value_col:
                        value_columns = [fallback_value_col]
                if not value_columns:
                    raise HTTPException(status_code=400, detail="No numeric value column found")

                if requested_source_code != "AUTO" and len(value_columns) > 1:
                    requested_source_code = "AUTO"
                    source_resolution_mode = "auto_by_column_multi_value"

                var_res = await db.execute(text("SELECT variable_id FROM ref.variable WHERE code = :code"), {"code": variable_code})
                var_row = var_res.first()
                if not var_row: raise HTTPException(404, f"Variable {variable_code} not found")
                var_id = var_row[0]
                var_family = _infer_variable_family(variable_code)
                if requested_source_code != "AUTO" and not _is_source_compatible_with_family(requested_source_code, var_family):
                    allowed = ", ".join(_allowed_sources_for_family(var_family))
                    raise HTTPException(
                        status_code=400,
                        detail=f"Source {requested_source_code} non compatible avec la variable {variable_code}. Sources autorisees: {allowed}",
                    )

                primary_quality_col = quality_columns[0] if quality_columns else None
                replaced_source_ids: Set[str] = set()

                for val_col in value_columns:
                    source_hint_for_col = _infer_source_code_from_column(val_col)
                    matching_quality_col = None
                    if source_hint_for_col:
                        matching_quality_col = next(
                            (
                                qcol for qcol in quality_columns
                                if _infer_source_code_from_column(qcol) == source_hint_for_col
                            ),
                            None,
                        )
                    if matching_quality_col is None:
                        matching_quality_col = primary_quality_col

                    _, effective_source_id, effective_run_id = await _resolve_source_payload(
                        db,
                        requested_source_code,
                        source_ids_map,
                        run_cache,
                        column_name=val_col,
                        variable_hint=variable_code,
                    )

                    values = []
                    for row in parsed_records:
                        val = _coerce_float(row.get(val_col))
                        if val is None:
                            continue
                        qc_flag = 0
                        if matching_quality_col:
                            qv = str(row.get(matching_quality_col) or "").strip().lower()
                            if qv in {"suspect", "warning", "douteux"}:
                                qc_flag = 1
                            elif qv in {"bad", "mauvais", "invalid", "rejete"}:
                                qc_flag = 2
                        ts_val = row[ts_col]
                        values.append(
                            {
                                "station_id": station_id,
                                "variable_id": var_id,
                                "time": ts_val,
                                "value": val,
                                "qc_flag": qc_flag,
                                "source_id": effective_source_id,
                                "run_id": effective_run_id,
                            }
                        )

                    if replace_existing.lower() == "true" and values:
                        source_key = str(effective_source_id)
                        if source_key not in replaced_source_ids:
                            min_ts = min(v["time"] for v in values)
                            max_ts = max(v["time"] for v in values)
                            if entity_type == "bassins":
                                await db.execute(
                                    text(
                                        """
                                        DELETE FROM ts.basin_measurement
                                        WHERE basin_id = CAST(:sid AS UUID) AND variable_id = :vid AND source_id = :src AND time BETWEEN :min_ts AND :max_ts
                                        """
                                    ),
                                    {"sid": station_id, "vid": var_id, "src": effective_source_id, "min_ts": min_ts, "max_ts": max_ts},
                                )
                            else:
                                await db.execute(
                                    text(
                                        """
                                        DELETE FROM ts.measurement
                                        WHERE station_id = CAST(:sid AS UUID) AND variable_id = :vid AND source_id = :src AND time BETWEEN :min_ts AND :max_ts
                                        """
                                    ),
                                    {"sid": station_id, "vid": var_id, "src": effective_source_id, "min_ts": min_ts, "max_ts": max_ts},
                                )
                            replaced_source_ids.add(source_key)

                    for v in values:
                        if entity_type == "bassins":
                            await db.execute(
                                text(
                                    """
                                    INSERT INTO ts.basin_measurement (basin_id, variable_id, time, value, source_id, run_id)
                                    VALUES (CAST(:station_id AS UUID), :variable_id, :time, :value, :source_id, :run_id)
                                    ON CONFLICT (time, basin_id, variable_id, source_id, COALESCE(run_id, '00000000-0000-0000-0000-000000000000'::uuid)) DO UPDATE SET value = EXCLUDED.value
                                    """
                                ),
                                {
                                    "station_id": v["station_id"],
                                    "variable_id": v["variable_id"],
                                    "time": v["time"],
                                    "value": v["value"],
                                    "source_id": v["source_id"],
                                    "run_id": v["run_id"],
                                },
                            )
                        else:
                            await db.execute(
                                text(
                                    """
                                    INSERT INTO ts.measurement (station_id, variable_id, time, value, qc_flag, source_id, run_id)
                                    VALUES (CAST(:station_id AS UUID), :variable_id, :time, :value, :qc_flag, :source_id, :run_id)
                                    ON CONFLICT (time, station_id, variable_id, source_id, COALESCE(run_id, '00000000-0000-0000-0000-000000000000'::uuid)) DO UPDATE SET value = EXCLUDED.value
                                    """
                                ),
                                v,
                            )
                    records_count += len(values)

            elif import_mode == "multi_station":
                if not variable_code: raise HTTPException(400, "Variable required for multi-station mode")
                
                # Check for Barrage constraint
                is_barrage_var = variable_code in BARRAGE_VARIABLE_CODES
                
                # Get var ID
                var_res = await db.execute(text("SELECT variable_id FROM ref.variable WHERE code = :code"), {"code": variable_code})
                var_row = var_res.first()
                if not var_row: raise HTTPException(404, f"Variable {variable_code} not found")
                var_id = var_row[0]
                var_family = _infer_variable_family(variable_code)
                if requested_source_code != "AUTO" and not _is_source_compatible_with_family(requested_source_code, var_family):
                    allowed = ", ".join(_allowed_sources_for_family(var_family))
                    raise HTTPException(
                        status_code=400,
                        detail=f"Source {requested_source_code} non compatible avec la variable {variable_code}. Sources autorisees: {allowed}",
                    )
                
                # Cache entity IDs
                entity_map = await _load_entity_map_safe(db, entity_type)
                try:
                    template_alias_map = _load_template_entity_alias_map(tmp_path, file.filename)
                except Exception as alias_exc:
                    print(f"Warning: template alias map failed ({alias_exc})")
                    template_alias_map = {}
                unmatched_columns: List[str] = []

                for col in columns:
                    if col == ts_col: continue
                    
                    if col.lower().startswith('unnamed'):
                        continue

                    st_info = None
                    for key in _alias_lookup_keys(col):
                        if key in entity_map:
                            st_info = entity_map[key]
                            break
                    if not st_info and template_alias_map:
                        alias_name = None
                        for key in _alias_lookup_keys(col):
                            alias_name = template_alias_map.get(key)
                            if alias_name:
                                break
                        if alias_name:
                            for alias_key in _entity_lookup_keys(alias_name):
                                if alias_key in entity_map:
                                    st_info = entity_map[alias_key]
                                    break
                    if not st_info:
                        print(f"Warning: Entity '{col}' not found in DB (normalized matching)")
                        unmatched_columns.append(col)
                        continue
                    
                    st_id = _entity_field(st_info, "id")
                    st_type = _entity_field(st_info, "type", "Station") or "Station"
                    if not st_id:
                        unmatched_columns.append(col)
                        continue
                    
                    # Validate Barrage constraint
                    if entity_type == "stations" and is_barrage_var and "barrage" not in str(st_type).lower():
                        print(f"Skipping station {col} for variable {variable_code} (Not a Barrage)")
                        continue

                    _, col_source_id, col_run_id = await _resolve_source_payload(
                        db,
                        requested_source_code,
                        source_ids_map,
                        run_cache,
                        column_name=col,
                        variable_hint=variable_code,
                    )

                    values = []
                    for row in parsed_records:
                        val = _coerce_float(row.get(col))
                        if val is not None:
                            ts_val = row[ts_col]
                            values.append({
                                "station_id": st_id,
                                "variable_id": var_id,
                                "time": ts_val,
                                "value": val,
                                "qc_flag": 0,
                                "source_id": col_source_id,
                                "run_id": col_run_id
                            })
                    print(f"Station {col}: Found {len(values)} valid values to insert.")
                    
                    if replace_existing.lower() == 'true' and values:
                        min_ts = min(v['time'] for v in values)
                        max_ts = max(v['time'] for v in values)
                        
                        # Execute DELETE
                        if entity_type == "bassins":
                            await db.execute(text("""
                                DELETE FROM ts.basin_measurement 
                                WHERE basin_id = CAST(:sid AS UUID) AND variable_id = :vid AND source_id = :src AND time BETWEEN :min_ts AND :max_ts
                            """), {"sid": st_id, "vid": var_id, "src": col_source_id, "min_ts": min_ts, "max_ts": max_ts})
                        else:
                            await db.execute(text("""
                                DELETE FROM ts.measurement 
                                WHERE station_id = CAST(:sid AS UUID) AND variable_id = :vid AND source_id = :src AND time BETWEEN :min_ts AND :max_ts
                            """), {"sid": st_id, "vid": var_id, "src": col_source_id, "min_ts": min_ts, "max_ts": max_ts})

                    for v in values:
                         if entity_type == "bassins":
                             await db.execute(text("""
                                INSERT INTO ts.basin_measurement (basin_id, variable_id, time, value, source_id, run_id)
                                VALUES (CAST(:station_id AS UUID), :variable_id, :time, :value, :source_id, :run_id)
                                ON CONFLICT (time, basin_id, variable_id, source_id, COALESCE(run_id, '00000000-0000-0000-0000-000000000000'::uuid)) DO UPDATE SET value = EXCLUDED.value
                             """), {"station_id": v["station_id"], "variable_id": v["variable_id"], "time": v["time"], "value": v["value"], "source_id": v["source_id"], "run_id": v["run_id"]})
                         else:
                             await db.execute(text("""
                                INSERT INTO ts.measurement (station_id, variable_id, time, value, qc_flag, source_id, run_id)
                                VALUES (CAST(:station_id AS UUID), :variable_id, :time, :value, :qc_flag, :source_id, :run_id)
                                ON CONFLICT (time, station_id, variable_id, source_id, COALESCE(run_id, '00000000-0000-0000-0000-000000000000'::uuid)) DO UPDATE SET value = EXCLUDED.value
                             """), v)
                    records_count += len(values)

            elif import_mode == "multi_variable":
                if not station_id: raise HTTPException(400, "Station required for multi-variable mode")
                
                st_type = ""
                if entity_type == "stations":
                    st_res = await db.execute(
                        text("SELECT station_type FROM geo.station WHERE station_id = CAST(:sid AS UUID)"),
                        {"sid": station_id},
                    )
                    st_row = st_res.first()
                    st_type = st_row[0] if st_row else ""

                variable_lookup = await _load_variable_lookup(db)
                
                for col in columns:
                    if col == ts_col or col.lower().startswith('unnamed'):
                        continue

                    resolved_var = _resolve_variable_from_column(col, variable_lookup)

                    if not resolved_var:
                        continue

                    var_id = resolved_var["variable_id"]
                    resolved_var_code = str(resolved_var["variable_code"])
                    var_family = _infer_variable_family(resolved_var_code)
                    
                    if (
                        entity_type == "stations"
                        and resolved_var_code in BARRAGE_VARIABLE_CODES
                        and "barrage" not in str(st_type).lower()
                    ):
                        continue

                    if requested_source_code != "AUTO" and not _is_source_compatible_with_family(requested_source_code, var_family):
                        policy_skipped_columns.append(
                            {
                                "column": col,
                                "variable_code": resolved_var_code,
                                "reason": f"Source {requested_source_code} incompatible avec famille {var_family or 'unknown'}",
                            }
                        )
                        continue

                    _, col_source_id, col_run_id = await _resolve_source_payload(
                        db,
                        requested_source_code,
                        source_ids_map,
                        run_cache,
                        column_name=col,
                        variable_hint=resolved_var_code,
                    )

                    values = []
                    for row in parsed_records:
                        val = _coerce_float(row.get(col))
                        if val is not None:
                            ts_val = row[ts_col]
                            values.append({
                                "station_id": station_id,
                                "variable_id": var_id,
                                "time": ts_val,
                                "value": val,
                                "qc_flag": 0,
                                "source_id": col_source_id,
                                "run_id": col_run_id
                            })
                             
                    if replace_existing.lower() == 'true' and values:
                        min_ts = min(v['time'] for v in values)
                        max_ts = max(v['time'] for v in values)
                        
                        if entity_type == "bassins":
                            await db.execute(text("""
                                DELETE FROM ts.basin_measurement 
                                WHERE basin_id = CAST(:sid AS UUID) AND variable_id = :vid AND source_id = :src AND time BETWEEN :min_ts AND :max_ts
                            """), {"sid": station_id, "vid": var_id, "src": col_source_id, "min_ts": min_ts, "max_ts": max_ts})
                        else:
                            await db.execute(text("""
                                DELETE FROM ts.measurement 
                                WHERE station_id = CAST(:sid AS UUID) AND variable_id = :vid AND source_id = :src AND time BETWEEN :min_ts AND :max_ts
                            """), {"sid": station_id, "vid": var_id, "src": col_source_id, "min_ts": min_ts, "max_ts": max_ts})

                    for v in values:
                        if entity_type == "bassins":
                            await db.execute(text("""
                                INSERT INTO ts.basin_measurement (basin_id, variable_id, time, value, source_id, run_id)
                                VALUES (CAST(:station_id AS UUID), :variable_id, :time, :value, :source_id, :run_id)
                                ON CONFLICT (time, basin_id, variable_id, source_id, COALESCE(run_id, '00000000-0000-0000-0000-000000000000'::uuid)) DO UPDATE SET value = EXCLUDED.value
                            """), {
                                "station_id": v["station_id"],
                                "variable_id": v["variable_id"],
                                "time": v["time"],
                                "value": v["value"],
                                "source_id": v["source_id"],
                                "run_id": v["run_id"],
                            })
                        else:
                            await db.execute(text("""
                                INSERT INTO ts.measurement (station_id, variable_id, time, value, qc_flag, source_id, run_id)
                                VALUES (CAST(:station_id AS UUID), :variable_id, :time, :value, :qc_flag, :source_id, :run_id)
                                ON CONFLICT (time, station_id, variable_id, source_id, COALESCE(run_id, '00000000-0000-0000-0000-000000000000'::uuid)) DO UPDATE SET value = EXCLUDED.value
                            """), v)
                    records_count += len(values)

            if records_count == 0:
                policy_details = ""
                if policy_skipped_columns:
                    policy_details = (
                        f" {len(policy_skipped_columns)} colonne(s) ignoree(s) car la source "
                        f"{requested_source_code} n'est pas compatible avec certaines variables."
                    )
                unmatched_details = ""
                if import_mode == "multi_station" and "unmatched_columns" in locals() and unmatched_columns:
                    preview = ", ".join(unmatched_columns[:8])
                    suffix = "..." if len(unmatched_columns) > 8 else ""
                    label = "bassins" if entity_type == "bassins" else "stations"
                    unmatched_details = (
                        f" {len(unmatched_columns)} {label} non reconnus: {preview}{suffix}."
                    )
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Aucun point importe. Verifiez le format du fichier, les noms/codes des "
                        f"colonnes (stations ou bassins), le timestamp et la source selectionnee.{policy_details}{unmatched_details}"
                    ),
                )
              
            await db.commit()
            response_payload: Dict[str, Any] = {
                "status": "success",
                "message": f"Imported {records_count} data points",
                "effective_source_code": requested_source_code,
                "source_resolution": source_resolution_mode,
                "detected_datetime_format": detected_dt_style,
                "detected_datetime_confidence": detected_dt_confidence,
                "timestamp_sanity": ts_sanity,
                "file_source_hint": file_source_hint,
                "file_source_hints": file_source_hints,
                "column_source_hints": column_source_hints,
            }
            if policy_skipped_columns:
                response_payload["skipped_columns_policy"] = policy_skipped_columns
                response_payload["skipped_columns_policy_count"] = len(policy_skipped_columns)
                response_payload["message"] = (
                    f"{response_payload['message']} | {len(policy_skipped_columns)} colonne(s) ignoree(s) "
                    "car source non compatible"
                )
            return response_payload

        except HTTPException:
            await db.rollback()
            raise
        except Exception as e:
            await db.rollback()
            import traceback
            traceback.print_exc()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            _safe_remove_tmp(tmp_path)
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Upload failed: {str(e)}")

# Request/Response models
class TimeSeriesPoint(BaseModel):
    timestamp: datetime
    value: float
    quality_flag: Optional[str] = None

class TimeSeriesCreate(BaseModel):
    timestamp: datetime
    value: float
    quality_flag: Optional[str] = "good"

@router.get("/timeseries/{variable_code}")
async def list_timeseries_stations(
    variable_code: str,
    all_stations: bool = Query(False),
    db: AsyncSession = Depends(get_db)
):
    """
    Get list of stations.
    If all_stations=True, returns ALL stations (for import).
    Else, returns only stations with data for this variable.
    """
    try:
        if all_stations:
             query = text("""
                SELECT 
                    s.station_id, 
                    s.code, 
                    s.name,
                    s.station_type,
                    0 as data_count
                FROM geo.station s
                ORDER BY s.name
            """)
             params = {}
        else:
            query = text("""
                SELECT DISTINCT 
                    s.station_id, 
                    s.code, 
                    s.name,
                    s.station_type,
                    COUNT(m.value) as data_count
                FROM geo.station s
                INNER JOIN ts.measurement m ON s.station_id = m.station_id
                INNER JOIN ref.variable v ON m.variable_id = v.variable_id
                WHERE v.code = :variable_code
                GROUP BY s.station_id, s.code, s.name, s.station_type
                ORDER BY s.name
            """)
            params = {"variable_code": variable_code}
        
        result = await db.execute(query, params)
        rows = result.mappings().all()
        
        return {
            "variable_code": variable_code,
            "stations": [dict(row) for row in rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching stations: {str(e)}")


@router.get("/timeseries/{variable_code}/{station_id}")
async def get_timeseries_data(
    variable_code: str,
    station_id: str,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Get time series data for a specific variable and station"""
    try:
        # Validate UUID
        safe_uuid = str(UUID(station_id))

        # Build query with optional date filters
        # INJECT UUID AS LITERAL STRING with CAST
        query_parts = [f"""
            SELECT 
                m.time as timestamp,
                m.value,
                m.qc_flag as quality_flag,
                v.code as variable_code,
                v.label as variable_name,
                v.unit
            FROM ts.measurement m
            INNER JOIN ref.variable v ON m.variable_id = v.variable_id
            WHERE m.station_id = CAST('{safe_uuid}' AS UUID) 
            AND v.code = :variable_code
        """]
        
        params = {
            "variable_code": variable_code
        }
        
        if start_date:
            query_parts.append("AND m.time >= :start_date")
            params["start_date"] = start_date
            
        if end_date:
            query_parts.append("AND m.time <= :end_date")
            params["end_date"] = end_date
            
        query_parts.append("ORDER BY m.time DESC LIMIT 1000")
        
        query = text(" ".join(query_parts))
        result = await db.execute(query, params)
        rows = result.mappings().all()
        
        return {
            "variable_code": variable_code,
            "station_id": station_id,
            "data_count": len(rows),
            "data": [dict(row) for row in rows]
        }
    except Exception as e:
        # Log error to stderr for capturing
        import sys
        print(f"ERROR in get_timeseries_data: {e}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Error fetching time series: {str(e)}")

@router.post("/timeseries/{variable_code}/{station_id}")
async def add_timeseries_point(
    variable_code: str,
    station_id: str,
    point: TimeSeriesCreate,
    db: AsyncSession = Depends(get_db)
):
    """Add a new time series data point"""
    try:
        safe_uuid = str(UUID(station_id))

        # Get variable_id
        var_query = text("SELECT variable_id FROM ref.variable WHERE code = :code")
        var_result = await db.execute(var_query, {"code": variable_code})
        var_row = var_result.first()
        
        if not var_row:
            raise HTTPException(status_code=404, detail=f"Variable {variable_code} not found")
        
        variable_id = var_row[0]
        
        # Insert measurement
        insert_query = text(f"""
            INSERT INTO ts.measurement (station_id, variable_id, time, value, qc_flag)
            VALUES (CAST('{safe_uuid}' AS UUID), :variable_id, :timestamp, :value, :quality_flag)
            RETURNING station_id
        """)
        
        result = await db.execute(insert_query, {
            "variable_id": variable_id,
            "timestamp": point.timestamp,
            "value": point.value,
            "quality_flag": point.quality_flag
        })
        
        await db.commit()
        
        return {
            "status": "success",
            "message": "Data point added"
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error adding data point: {str(e)}")

@router.delete("/timeseries/{variable_code}/{station_id}/{timestamp}")
async def delete_timeseries_point(
    variable_code: str,
    station_id: str,
    timestamp: str,
    db: AsyncSession = Depends(get_db)
):
    """Delete a single time series data point"""
    try:
        safe_uuid = str(UUID(station_id))
        
        from datetime import datetime
        try:
            ts_datetime = datetime.fromisoformat(timestamp)
        except ValueError:
            # Fallback if timestamp has unexpected format
            ts_datetime = timestamp

        query = text(f"""
            DELETE FROM ts.measurement m
            USING ref.variable v
            WHERE m.station_id = CAST('{station_id}' AS UUID)
            AND m.variable_id = v.variable_id
            AND v.code = :variable_code
            AND m.time = :timestamp
            RETURNING m.station_id
        """)
        result = await db.execute(query, {"variable_code": variable_code, "timestamp": ts_datetime})
        
        deleted = result.first()
        
        if not deleted:
            raise HTTPException(status_code=404, detail="Measurement not found")
        
        await db.commit()
        
        return {
            "status": "success",
            "message": "Data point deleted"
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting data point: {str(e)}")


@router.delete("/timeseries/{variable_code}/{station_id}")
async def delete_timeseries_series(
    variable_code: str,
    station_id: str,
    db: AsyncSession = Depends(get_db)
):
    """Delete ALL time series data for a specific variable and station"""
    try:
        safe_uuid = str(UUID(station_id))
        
        # Get variable_id first
        var_query = text("SELECT variable_id FROM ref.variable WHERE code = :code")
        var_result = await db.execute(var_query, {"code": variable_code})
        var_row = var_result.first()
        
        if not var_row:
            raise HTTPException(status_code=404, detail=f"Variable {variable_code} not found")
        
        variable_id = var_row[0]
        
        query = text(f"""
            DELETE FROM ts.measurement
            WHERE station_id = CAST('{safe_uuid}' AS UUID)
            AND variable_id = :variable_id
        """)
        
        result = await db.execute(query, {"variable_id": variable_id})
        deleted_count = result.rowcount
        
        await db.commit()
        
        return {
            "status": "success",
            "message": f"Deleted {deleted_count} measurements for variable {variable_code}"
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting time series: {str(e)}")
