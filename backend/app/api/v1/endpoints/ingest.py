import os
import shutil
import subprocess
import tempfile
import re
import unicodedata
from datetime import datetime
from typing import List, Optional, Any, Dict
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from pydantic import BaseModel
import psycopg
from psycopg.rows import dict_row
from app.core.config import settings

router = APIRouter()

# Path calculation
# current file: backend/app/api/v1/endpoints/ingest.py
# target: backend/../Data/scripts (relative to backend root)
# Let's assume we can find it relative to this file
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))
# BASE_DIR should be .../backend
# Actually, let's just go up to 'hydro_sentinel' root
PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, ".."))
SCRIPTS_DIR = os.path.join(PROJECT_ROOT, "Data", "scripts")

# Check if scripts dir exists, if not try another path or fail gracefully
if not os.path.exists(SCRIPTS_DIR):
    # Fallback/Debug: maybe we are in a different structure?
    # backend is c:\dev\detection_inondation\hydro_sentinel\backend
    # scripts are c:\dev\detection_inondation\hydro_sentinel\Data\scripts
    SCRIPTS_DIR = r"c:\dev\detection_inondation\hydro_sentinel\Data\scripts"


def _normalize_obs_key(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower()
    text = text.replace("(mm)", "")
    text = text.replace("pluie", "")
    text = text.replace("1hr", "")
    text = text.replace("1h", "")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _looks_like_mapping_header(header_cells: List[str]) -> bool:
    joined = " | ".join(header_cells)
    has_code = "code" in joined
    has_station = ("nom station" in joined) or ("station nom" in joined) or ("name" in joined)
    has_external = ("fichier" in joined) or ("pcp" in joined) or ("obs" in joined)
    return has_code and has_station and has_external


def _load_precip_obs_mapping_from_workbook(excel_path: str) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return mapping

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(min_row=1, max_row=40, values_only=True)
            all_rows = [list(r or []) for r in rows]
            header_idx: Optional[int] = None
            code_idx: Optional[int] = None
            station_idx: Optional[int] = None
            external_idx: Optional[int] = None

            for idx, row in enumerate(all_rows):
                normalized = [str(c or "").strip().lower() for c in row]
                if not _looks_like_mapping_header(normalized):
                    continue
                header_idx = idx
                for cidx, col in enumerate(normalized):
                    if col == "code" or "station_code" in col:
                        code_idx = cidx
                    if col in {"nom station", "station nom", "name", "nom"} or "nom station" in col:
                        station_idx = cidx
                    if (
                        "nom station fichier" in col
                        or "station fichier" in col
                        or "pcp" in col
                        or "fichier" in col
                    ):
                        external_idx = cidx
                break

            if header_idx is None:
                continue

            if station_idx is None:
                station_idx = 1
            if code_idx is None:
                code_idx = 0

            # If external column is not explicitly found, use code as fallback key.
            for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
                row_list = list(row or [])
                if all(str(c or "").strip() == "" for c in row_list):
                    continue
                station_name = str(row_list[station_idx] if station_idx < len(row_list) else "").strip()
                if not station_name:
                    continue
                code_val = str(row_list[code_idx] if code_idx < len(row_list) else "").strip()
                external_val = (
                    str(row_list[external_idx]).strip()
                    if external_idx is not None and external_idx < len(row_list)
                    else ""
                )
                for candidate in (external_val, code_val):
                    key = _normalize_obs_key(candidate)
                    if key:
                        mapping.setdefault(key, station_name)
    finally:
        wb.close()

    return mapping


def _apply_precip_obs_mapping_to_workbook(excel_path: str) -> int:
    mapping = _load_precip_obs_mapping_from_workbook(excel_path)
    if not mapping:
        return 0

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return 0

    wb = load_workbook(excel_path)
    renamed = 0
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=30):
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    key = _normalize_obs_key(value)
                    target = mapping.get(key)
                    if target and target != value:
                        cell.value = target
                        renamed += 1
        if renamed > 0:
            wb.save(excel_path)
    finally:
        wb.close()

    return renamed

class IngestionResponse(BaseModel):
    status: str
    message: str
    logs: str
    ingestion_id: Optional[str] = None

@router.get("/history", response_model=List[Any])
def get_ingestions(limit: int = 50):
    try:
        dsn = str(settings.DATABASE_URL)
        # SQLAlchemy URI might look like postgresql+psycopg2://...
        # psycopg needs postgresql://...
        if "+psycopg2" in dsn:
            dsn = dsn.replace("+psycopg2", "")
        if "+asyncpg" in dsn:
            dsn = dsn.replace("+asyncpg", "")
            
        with psycopg.connect(dsn, row_factory=dict_row) as conn:
            res = conn.execute("""
                SELECT ingestion_id, status, pipeline_name, started_at, finished_at, summary 
                FROM files.ingestion 
                ORDER BY started_at DESC 
                LIMIT %s
            """, (limit,)).fetchall()
            return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/analyze", response_model=Dict[str, Any])
async def analyze_ingestion(
    file: UploadFile = File(...),
    type: str = Form(...), # "abhs", "precip", "datatable"
    source_code: str = Form("ABHS_RES"),
    run_date: str = Form(None)
):
    from app.services.ingestion_analyzer import IngestionAnalyzer
    
    analyzer = IngestionAnalyzer()
    content = await file.read()
    
    # Map frontend type to analyzer type
    # Frontend sends: "abhs" (Results), "precip" (Rainfall), "datatable" (Stations)
    report = await analyzer.analyze_file(content, file.filename, type)
    
    # Reset file cursor for potential further use or just close
    await file.seek(0)
    
    return report

@router.post("/execute", response_model=IngestionResponse)
async def execute_ingestion(
    file: UploadFile = File(...),
    type: str = Form(...),
    source_code: str = Form("ABHS_RES"),
    run_date: str = Form(None)
):
    return await run_ingestion(file, type, source_code, run_date, dry_run=False)

async def run_ingestion(file: UploadFile, type: str, source_code: str, run_date: str, dry_run: bool):
    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        dsn = str(settings.DATABASE_URL)
        if "+psycopg2" in dsn:
            dsn = dsn.replace("+psycopg2", "")
        if "+asyncpg" in dsn:
            dsn = dsn.replace("+asyncpg", "")

        cmd = []
        renamed_count = 0
        if type == "abhs":
            script = os.path.join(SCRIPTS_DIR, "ingest_excel_abhs_results.py")
            if not run_date:
                run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            cmd = [
                "python", script,
                "--dsn", dsn,
                "--excel", tmp_path,
                "--source-code", source_code,
                "--run-time", run_date,
                "--run-label", f"Import via Web {datetime.now().strftime('%H:%M')}"
            ]
        elif type == "precip":
            script = os.path.join(SCRIPTS_DIR, "ingest_excel_precip_obs_v1.2.py")
            renamed_count = _apply_precip_obs_mapping_to_workbook(tmp_path)
            cmd = [
                "python", script,
                "--dsn", dsn,
                "--excel", tmp_path
            ]
        else:
            raise HTTPException(status_code=400, detail="Invalid ingestion type")

        if dry_run:
            cmd.append("--dry-run")
            
        # Ensure script exists
        if not os.path.exists(script):
             return {
                "status": "error",
                "message": f"Script not found at {script}",
                "logs": f"Checked path: {script}"
            }

        # Run script
        # Capture environment to pass potentially needed vars
        env = os.environ.copy()
        
        process = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            env=env
        )

        logs = process.stdout + "\n" + process.stderr
        if type == "precip":
            logs = f"[OBS mapping] colonnes renommees automatiquement: {renamed_count}\n" + logs
        status = "success" if process.returncode == 0 else "error"
        
        message = "Ingestion analysis complete" if dry_run else "Ingestion complete"
        if process.returncode != 0:
            message = "Ingestion script failed"

        return {
            "status": status,
            "message": message,
            "logs": logs
        }

    except Exception as e:
        return {
            "status": "error",
            "message": f"Server error: {str(e)}",
            "logs": str(e)
        }
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
