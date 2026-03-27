from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends, Query, Response
from typing import List, Optional, Any, Dict
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from app.core.config import settings
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from app.db.session import get_db
import os
import sys
import shutil
import tempfile
import zipfile
import json
import csv
import io
import re
from datetime import datetime, timedelta
from pathlib import Path



router = APIRouter()


def _ensure_openpyxl_available() -> None:
    """
    Ensure openpyxl is importable for template generation.
    Falls back to common Conda site-packages paths on Windows when backend
    is started with a different Python interpreter.
    """
    try:
        import openpyxl  # noqa: F401
        return
    except ModuleNotFoundError:
        pass

    fallback_paths = []
    conda_prefix = os.environ.get("CONDA_PREFIX")
    if conda_prefix:
        fallback_paths.append(os.path.join(conda_prefix, "Lib", "site-packages"))

    if os.name == "nt":
        fallback_paths.append(r"C:\anaconda\Lib\site-packages")
        fallback_paths.append(r"C:\ProgramData\anaconda3\Lib\site-packages")

    for path in fallback_paths:
        if path and os.path.isdir(path) and path not in sys.path:
            sys.path.append(path)
            try:
                import openpyxl  # noqa: F401
                return
            except ModuleNotFoundError:
                continue

    raise HTTPException(
        status_code=500,
        detail="openpyxl is required for Excel templates. Install it in the backend environment.",
    )


# --- Models ---
class EntityUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    station_type: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None
    attributes: Optional[Dict[str, Any]] = None

class EntityCreate(BaseModel):
    name: str
    code: str
    station_type: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None

# --- Helper ---
ALLOWED_TABLES = {
    "stations": "geo.station",
    "bassins": "geo.basin",
}

def get_table(entity_type: str) -> str:
    table = ALLOWED_TABLES.get(entity_type)
    if not table:
        raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}. Allowed: stations, bassins")
    return table

# --- LIST ---
@router.get("/entities/{entity_type}")
async def list_entities(entity_type: str, db: AsyncSession = Depends(get_db)):
    """List entities like 'stations', 'bassins' from DB geo tables"""
    table_map = {
        "stations": "geo.station",
        "bassins": "geo.basin",
        "barrages": "geo.barrage"
    }
    
    target_table = table_map.get(entity_type)
    if not target_table:
        target_table = f"geo.{entity_type}"

    try:
        allowed_tables = ["geo.station", "geo.basin", "geo.barrage", "ref.station_alias", "ref.basin_alias"]
        if target_table not in allowed_tables:
            raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")

        result = await db.execute(text(f"SELECT * FROM {target_table} LIMIT 500"))
        rows = result.mappings().all()
        return [dict(row) for row in rows]
    except Exception as e:
        error_msg = str(e).lower()
        if "n'existe pas" in error_msg or "does not exist" in error_msg or "relation" in error_msg:
            return []
        raise HTTPException(status_code=500, detail=str(e))


# --- CREATE ---
@router.post("/entities/{entity_type}")
async def create_entity(entity_type: str, entity: EntityCreate, db: AsyncSession = Depends(get_db)):
    """Create a new station or bassin"""
    target_table = get_table(entity_type)
    
    try:
        if entity_type == "stations":
            # Enforce coordinates for stations as geom is NOT NULL
            if entity.lat is None or entity.lon is None:
                 raise HTTPException(status_code=400, detail="Latitude and Longitude are required for new stations")

            stmt = text(f"""
                INSERT INTO {target_table} (name, code, station_type, geom)
                VALUES (:name, :code, :station_type, ST_SetSRID(ST_MakePoint(:lon, :lat), 4326))
                RETURNING station_id, name, code, station_type
            """)
            result = await db.execute(stmt, {
                "name": entity.name,
                "code": entity.code,
                "station_type": entity.station_type or "Station hydrologique",
                "lat": entity.lat,
                "lon": entity.lon
            })
        else:  # bassins
            if entity.lat is not None and entity.lon is not None:
                stmt = text(f"""
                    INSERT INTO {target_table} (name, code, geom)
                    VALUES (:name, :code, ST_SetSRID(ST_MakePoint(:lon, :lat), 4326))
                    RETURNING basin_id, name, code
                """)
                result = await db.execute(stmt, {
                    "name": entity.name,
                    "code": entity.code,
                    "lat": entity.lat,
                    "lon": entity.lon
                })
            else:
                stmt = text(f"""
                    INSERT INTO {target_table} (name, code)
                    VALUES (:name, :code)
                    RETURNING basin_id, name, code
                """)
                result = await db.execute(stmt, {
                    "name": entity.name,
                    "code": entity.code
                })
        
        row = result.mappings().first()
        await db.commit()
        return {"status": "success", "entity": dict(row) if row else {}}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating entity: {str(e)}")


# --- UPDATE ---
@router.put("/entities/{entity_type}/{entity_id}")
async def update_entity(entity_type: str, entity_id: str, entity: EntityUpdate, db: AsyncSession = Depends(get_db)):
    """Update a station or bassin"""
    target_table = get_table(entity_type)
    
    try:
        # Build SET clause dynamically
        set_parts = []
        params: Dict[str, Any] = {"entity_id": entity_id}
        
        if entity.name is not None:
            set_parts.append("name = :name")
            params["name"] = entity.name
        if entity.code is not None:
            set_parts.append("code = :code")
            params["code"] = entity.code
        if entity_type == "stations" and entity.station_type is not None:
            set_parts.append("station_type = :station_type")
            params["station_type"] = entity.station_type
        if entity.lat is not None and entity.lon is not None:
            set_parts.append("geom = ST_SetSRID(ST_MakePoint(:lon, :lat), 4326)")
            params["lat"] = entity.lat
            params["lon"] = entity.lon
        
        if not set_parts:
            raise HTTPException(status_code=400, detail="No fields to update")
        
        # Determine ID column
        id_col = "station_id" if entity_type == "stations" else "basin_id"
        
        stmt = text(f"""
            UPDATE {target_table}
            SET {', '.join(set_parts)}
            WHERE {id_col} = CAST(:entity_id AS UUID)
            RETURNING {id_col}, name, code
        """)
        
        result = await db.execute(stmt, params)
        row = result.mappings().first()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"Entity {entity_id} not found")
        
        await db.commit()
        return {"status": "success", "entity": dict(row)}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating entity: {str(e)}")


# --- DELETE ---
@router.delete("/entities/{entity_type}/{entity_id}")
async def delete_entity(entity_type: str, entity_id: str, db: AsyncSession = Depends(get_db)):
    """Delete a station or bassin (with cascade on measurements)"""
    target_table = get_table(entity_type)
    id_col = "station_id" if entity_type == "stations" else "basin_id"
    
    try:
        # For stations: delete measurements first
        if entity_type == "stations":
            await db.execute(
                text(f"DELETE FROM ts.measurement WHERE station_id = CAST(:id AS UUID)"),
                {"id": entity_id}
            )
            
            # Also cleanup dss_mapping which references station_id
            try:
                await db.execute(
                    text(f"DELETE FROM dss_mapping WHERE station_id = CAST(:id AS UUID)"),
                    {"id": entity_id}
                )
            except Exception:
                pass # Table might not exist or error, proceed

            # Delete from dss_mapping if exists
            try:
                await db.execute(
                    text(f"DELETE FROM dss_mapping WHERE station_id = CAST(:id AS UUID)"),
                    {"id": entity_id}
                )
            except Exception:
                pass # Table might not exist or other error, but proceed

        
        result = await db.execute(
            text(f"DELETE FROM {target_table} WHERE {id_col} = CAST(:id AS UUID) RETURNING {id_col}"),
            {"id": entity_id}
        )
        deleted = result.first()
        
        if not deleted:
            raise HTTPException(status_code=404, detail=f"Entity {entity_id} not found")
        
        await db.commit()
        return {"status": "success", "message": f"Entity {entity_id} deleted"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting entity: {str(e)}")


# --- TEMPLATE GENERATION ---

TEMPLATE_ALLOWED_SOURCE_CODES = {"OBS", "AROME", "ECMWF", "SIM"}
TEMPLATE_SOURCE_LABELS = {
    "OBS": "Observations",
    "AROME": "Prevision AROME",
    "ECMWF": "Prevision ECMWF",
    "SIM": "Simule",
}
BASIN_SHAPE_ALLOWED = {"ABH", "DGM"}


def _normalize_template_source_code(source_code: Optional[str]) -> str:
    normalized = (source_code or "OBS").strip().upper()
    if normalized == "AUTO":
        return "OBS"
    if normalized not in TEMPLATE_ALLOWED_SOURCE_CODES:
        allowed = ", ".join(sorted(TEMPLATE_ALLOWED_SOURCE_CODES))
        raise HTTPException(status_code=400, detail=f"Invalid source_code: {normalized}. Allowed: {allowed}")
    return normalized


def _normalize_template_source_codes(source_codes: Optional[str]) -> List[str]:
    raw = str(source_codes or "").strip()
    if not raw:
        return ["OBS"]

    parts = [p.strip().upper() for p in re.split(r"[,\s;|]+", raw) if p and p.strip()]
    normalized_codes: List[str] = []
    for part in parts:
        if part == "AUTO":
            part = "OBS"
        if part not in TEMPLATE_ALLOWED_SOURCE_CODES:
            allowed = ", ".join(sorted(TEMPLATE_ALLOWED_SOURCE_CODES))
            raise HTTPException(status_code=400, detail=f"Invalid source code: {part}. Allowed: {allowed}")
        if part not in normalized_codes:
            normalized_codes.append(part)

    return normalized_codes or ["OBS"]


def _normalize_basin_shape(shape: Optional[str]) -> str:
    normalized = (shape or "ABH").strip().upper()
    if normalized not in BASIN_SHAPE_ALLOWED:
        allowed = ", ".join(sorted(BASIN_SHAPE_ALLOWED))
        raise HTTPException(status_code=400, detail=f"Invalid basin_shape: {normalized}. Allowed: {allowed}")
    return normalized


def _load_dgm_basins_from_geojson() -> List[Dict[str, str]]:
    # Try common project locations for local DGM GeoJSON.
    root = Path(__file__).resolve().parents[5]
    candidates = [
        root / "hydro-sentinel" / "public" / "data" / "basins_dgm.geojson",
        root / "public" / "data" / "basins_dgm.geojson",
    ]
    for path in candidates:
        if not path.exists():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            features = payload.get("features") if isinstance(payload, dict) else []
            if not isinstance(features, list):
                continue
            entities: List[Dict[str, str]] = []
            for idx, feature in enumerate(features):
                props = feature.get("properties", {}) if isinstance(feature, dict) else {}
                name = (
                    props.get("name")
                    or props.get("nom")
                    or props.get("NOM")
                    or props.get("Name")
                    or props.get("Name1")
                    or props.get("BASSIN")
                    or f"Bassin DGM {idx + 1}"
                )
                code = props.get("code") or props.get("CODE") or props.get("Code") or props.get("id") or props.get("ID")
                entities.append(
                    {
                        "entity_id": f"dgm-{idx + 1}",
                        "name": str(name),
                        "code": str(code) if code is not None else f"DGM-{idx + 1}",
                    }
                )
            if entities:
                return entities
        except Exception:
            continue
    return []


def _source_suffix(source_code: str) -> str:
    return source_code.strip().lower()


def _append_source_suffix(value: Optional[str], source_code: str) -> str:
    base = str(value or "").strip()
    if not base:
        return base
    suffix = _source_suffix(source_code)
    normalized = base.lower()
    if normalized.endswith(f"_{suffix}"):
        return base
    return f"{base}_{suffix}"


def _sources_suffix(source_codes: List[str]) -> str:
    return "-".join(_source_suffix(code) for code in source_codes)


def _format_source_summary(source_codes: List[str]) -> str:
    return ", ".join(
        f"{TEMPLATE_SOURCE_LABELS.get(code, code)} ({code})"
        for code in source_codes
    )

@router.get("/templates/simple")
async def get_template_simple(
    station_id: Optional[str] = Query(None),
    variable_code: Optional[str] = Query(None),
    source_code: Optional[str] = Query("OBS"),
    db: AsyncSession = Depends(get_db)
):
    """Generate a pre-filled simple Excel template (one station, one variable)"""
    _ensure_openpyxl_available()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    normalized_source = _normalize_template_source_code(source_code)
    source_label = TEMPLATE_SOURCE_LABELS.get(normalized_source, normalized_source)
    
    # Get station info
    station_name = "STATION_CODE"
    variable_label = variable_code or "VARIABLE"
    
    if station_id:
        try:
            res = await db.execute(
                text("SELECT name, code FROM geo.station WHERE station_id = CAST(:id AS UUID)"),
                {"id": station_id}
            )
            row = res.mappings().first()
            if row:
                station_name = row["code"] or row["name"]
        except:
            pass
    
    if variable_code:
        try:
            res = await db.execute(
                text("SELECT label, unit FROM ref.variable WHERE code = :code"),
                {"code": variable_code}
            )
            row = res.mappings().first()
            if row:
                variable_label = f"{row['label']} ({row['unit']})"
        except:
            pass

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DonnÃ©es"
    
    # Style
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    info_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    
    # Info row
    ws["A1"] = f"Station: {station_name} | Variable: {variable_label} | Source: {source_label} ({normalized_source})"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].fill = info_fill
    ws.merge_cells("A1:C1")    ws["A2"] = (
        "Format: timestamp ISO8601, valeur numerique, flag qualite (good/suspect/bad) | "
        f"colonne valeur suffixee: _{_source_suffix(normalized_source)}"
    )
    ws["A2"].fill = info_fill
    ws.merge_cells("A2:C2")
    
    # Header row
    headers = ["timestamp", _append_source_suffix("value", normalized_source), "quality_flag"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=3, column=i+1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22 if i == 0 else 15
    
    # Example rows
    base_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(3):
        dt = base_date - timedelta(days=i)
        ws.append([dt.strftime("%Y-%m-%dT%H:%M:%S"), "", "good"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"template_simple_{station_name}_{variable_code or 'variable'}_{normalized_source.lower()}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/templates/multi-variable")
async def get_template_multi_variable(
    station_id: Optional[str] = Query(None),
    source_code: Optional[str] = Query("OBS"),
    db: AsyncSession = Depends(get_db)
):
    """Generate Excel template with columns = variables (for one station)"""
    _ensure_openpyxl_available()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    normalized_source = _normalize_template_source_code(source_code)
    source_label = TEMPLATE_SOURCE_LABELS.get(normalized_source, normalized_source)

    station_name = "STATION"
    station_code = "CODE"

    # Get station info
    if station_id:
        try:
            res = await db.execute(
                text("SELECT name, code FROM geo.station WHERE station_id = CAST(:id AS UUID)"),
                {"id": station_id}
            )
            row = res.mappings().first()
            if row:
                station_name = row["name"]
                station_code = row["code"] or row["name"]
        except:
            pass

    # Get all variables
    variables = []
    try:
        res = await db.execute(text("SELECT code, label, unit FROM ref.variable ORDER BY label"))
        variables = [dict(r) for r in res.mappings().all()]
    except:
        variables = [{"code": "precip_mm", "label": "Precipitations", "unit": "mm"}]

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donnees"

    # Style
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    info_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    # Info row
    ws["A1"] = f"Station: {station_name} ({station_code}) | Source: {source_label} ({normalized_source})"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].fill = info_fill
    ws.merge_cells(f"A1:{get_column_letter(len(variables) + 1)}1")

    ws["A2"] = (
        "Format: timestamp ISO8601 (YYYY-MM-DDTHH:MM:SS), valeurs numeriques par colonne variable | "
        f"suffixe source: _{_source_suffix(normalized_source)}"
    )
    ws["A2"].fill = info_fill
    ws.merge_cells(f"A2:{get_column_letter(len(variables) + 1)}2")

    # Header row
    ws["A3"] = "timestamp"
    ws["A3"].font = header_font
    ws["A3"].fill = header_fill
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22

    for i, var in enumerate(variables):
        col = get_column_letter(i + 2)
        cell = ws[f"{col}3"]
        source_code_col = _append_source_suffix(str(var["code"]), normalized_source)
        cell.value = f"{source_code_col}\n({var['unit']})"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[col].width = 16

    ws.row_dimensions[3].height = 35

    # Example rows
    base_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(3):
        dt = base_date - timedelta(days=i)
        row_num = 4 + i
        ws[f"A{row_num}"] = dt.strftime("%Y-%m-%dT%H:%M:%S")
        for j in range(len(variables)):
            ws[f"{get_column_letter(j + 2)}{row_num}"] = ""

    # Info sheet
    ws_info = wb.create_sheet("Variables")
    ws_info["A1"] = "Code"
    ws_info["B1"] = "Label"
    ws_info["C1"] = "Unite"
    for i, var in enumerate(variables):
        ws_info[f"A{i+2}"] = _append_source_suffix(str(var["code"]), normalized_source)
        ws_info[f"B{i+2}"] = var["label"]
        ws_info[f"C{i+2}"] = var["unit"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"template_multi_variable_{station_code}_{normalized_source.lower()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/templates/simple-multi-source")
async def get_template_simple_multi_source(
    station_id: Optional[str] = Query(None),
    variable_code: Optional[str] = Query(None),
    source_codes: Optional[str] = Query("OBS,SIM"),
    db: AsyncSession = Depends(get_db),
):
    """Generate a pre-filled simple template with multiple source columns."""
    _ensure_openpyxl_available()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    normalized_sources = _normalize_template_source_codes(source_codes)
    source_summary = _format_source_summary(normalized_sources)

    station_name = "STATION_CODE"
    variable_label = variable_code or "VARIABLE"

    if station_id:
        try:
            res = await db.execute(
                text("SELECT name, code FROM geo.station WHERE station_id = CAST(:id AS UUID)"),
                {"id": station_id},
            )
            row = res.mappings().first()
            if row:
                station_name = row["code"] or row["name"]
        except Exception:
            pass

    if variable_code:
        try:
            res = await db.execute(
                text("SELECT label, unit FROM ref.variable WHERE code = :code"),
                {"code": variable_code},
            )
            row = res.mappings().first()
            if row:
                variable_label = f"{row['label']} ({row['unit']})"
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donnees"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    info_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    headers = ["timestamp"]
    for code in normalized_sources:
        headers.append(_append_source_suffix("value", code))
        headers.append(_append_source_suffix("quality_flag", code))

    max_col = get_column_letter(len(headers))
    ws["A1"] = f"Station: {station_name} | Variable: {variable_label} | Sources: {source_summary}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].fill = info_fill
    ws.merge_cells(f"A1:{max_col}1")

    ws["A2"] = (
        "Format: timestamp ISO8601, colonnes valeur et qualite par source (good/suspect/bad) | "
        f"sources: {', '.join(normalized_sources)}"
    )
    ws["A2"].fill = info_fill
    ws.merge_cells(f"A2:{max_col}2")

    for i, header in enumerate(headers):
        cell = ws.cell(row=3, column=i + 1, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22 if i == 0 else 16

    base_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(3):
        dt = base_date - timedelta(days=i)
        row_values: List[Any] = [dt.strftime("%Y-%m-%dT%H:%M:%S")]
        for _ in normalized_sources:
            row_values.extend(["", "good"])
        ws.append(row_values)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        f"template_simple_multi_source_{station_name}_{variable_code or 'variable'}_"
        f"{_sources_suffix(normalized_sources)}.xlsx"
    )
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/templates/multi-variable-multi-source")
async def get_template_multi_variable_multi_source(
    station_id: Optional[str] = Query(None),
    source_codes: Optional[str] = Query("OBS,SIM"),
    db: AsyncSession = Depends(get_db),
):
    """Generate a multi-variable template with one column per variable/source pair."""
    _ensure_openpyxl_available()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    normalized_sources = _normalize_template_source_codes(source_codes)
    source_summary = _format_source_summary(normalized_sources)

    station_name = "STATION"
    station_code = "CODE"

    if station_id:
        try:
            res = await db.execute(
                text("SELECT name, code FROM geo.station WHERE station_id = CAST(:id AS UUID)"),
                {"id": station_id},
            )
            row = res.mappings().first()
            if row:
                station_name = row["name"]
                station_code = row["code"] or row["name"]
        except Exception:
            pass

    variables: List[Dict[str, Any]] = []
    try:
        res = await db.execute(text("SELECT code, label, unit FROM ref.variable ORDER BY label"))
        variables = [dict(r) for r in res.mappings().all()]
    except Exception:
        variables = [{"code": "precip_mm", "label": "Precipitations", "unit": "mm"}]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donnees"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    info_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    total_value_columns = max(1, len(variables) * len(normalized_sources))
    max_col = get_column_letter(total_value_columns + 1)
    ws["A1"] = f"Station: {station_name} ({station_code}) | Sources: {source_summary}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].fill = info_fill
    ws.merge_cells(f"A1:{max_col}1")

    ws["A2"] = (
        "Format: timestamp ISO8601 (YYYY-MM-DDTHH:MM:SS), valeurs numeriques par colonne variable_source "
        f"(ex: flow_m3s_obs) | sources: {', '.join(normalized_sources)}"
    )
    ws["A2"].fill = info_fill
    ws.merge_cells(f"A2:{max_col}2")

    ws["A3"] = "timestamp"
    ws["A3"].font = header_font
    ws["A3"].fill = header_fill
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22

    col_idx = 2
    for var in variables:
        for source in normalized_sources:
            col_letter = get_column_letter(col_idx)
            var_with_source = _append_source_suffix(str(var["code"]), source)
            cell = ws[f"{col_letter}3"]
            cell.value = f"{var_with_source}\n({var['unit']})"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[col_letter].width = 18
            col_idx += 1
    ws.row_dimensions[3].height = 40

    base_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(3):
        row_num = 4 + i
        ws[f"A{row_num}"] = (base_date - timedelta(days=i)).strftime("%Y-%m-%dT%H:%M:%S")
        for c in range(2, col_idx):
            ws[f"{get_column_letter(c)}{row_num}"] = ""

    ws_info = wb.create_sheet("Variables")
    ws_info["A1"] = "Code"
    ws_info["B1"] = "Source"
    ws_info["C1"] = "Label"
    ws_info["D1"] = "Unite"
    ws_info["E1"] = "Colonne"

    info_row = 2
    for var in variables:
        for source in normalized_sources:
            code_with_source = _append_source_suffix(str(var["code"]), source)
            ws_info[f"A{info_row}"] = str(var["code"])
            ws_info[f"B{info_row}"] = source
            ws_info[f"C{info_row}"] = var["label"]
            ws_info[f"D{info_row}"] = var["unit"]
            ws_info[f"E{info_row}"] = code_with_source
            info_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        f"template_multi_variable_multi_source_{station_code}_"
        f"{_sources_suffix(normalized_sources)}.xlsx"
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/templates/multi-station")
async def get_template_multi_station(
    variable_code: Optional[str] = Query(None),
    entity_type: str = Query("stations"),
    source_code: Optional[str] = Query("OBS"),
    basin_shape: Optional[str] = Query("ABH"),
    db: AsyncSession = Depends(get_db)
):
    """Generate Excel template with columns = stations or bassins (for one variable)."""
    _ensure_openpyxl_available()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    normalized_source = _normalize_template_source_code(source_code)
    source_label = TEMPLATE_SOURCE_LABELS.get(normalized_source, normalized_source)
    normalized_basin_shape = _normalize_basin_shape(basin_shape)

    normalized_entity_type = (entity_type or "stations").lower()
    if normalized_entity_type == "basins":
        normalized_entity_type = "bassins"
    if normalized_entity_type not in ["stations", "bassins"]:
        raise HTTPException(status_code=400, detail="Invalid entity_type. Allowed: stations, bassins")

    is_basin_template = normalized_entity_type == "bassins"
    entity_label_singular = "bassin" if is_basin_template else "station"
    entity_label_plural = "bassins" if is_basin_template else "stations"

    variable_label = variable_code or "VARIABLE"
    variable_unit = ""

    # Get variable info
    if variable_code:
        try:
            res = await db.execute(
                text("SELECT label, unit FROM ref.variable WHERE code = :code"),
                {"code": variable_code}
            )
            row = res.mappings().first()
            if row:
                variable_label = row["label"]
                variable_unit = row["unit"]
        except:
            pass

    # Get all entities
    entities = []
    try:
        if is_basin_template:
            provider_column_check = await db.execute(
                text(
                    """
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_schema = 'geo'
                      AND table_name = 'basin'
                      AND column_name = 'provider'
                    LIMIT 1
                    """
                )
            )
            provider_column_exists = provider_column_check.first() is not None

            if normalized_basin_shape == "DGM":
                # Always prefer local DGM shape file to guarantee expected basin naming.
                entities = _load_dgm_basins_from_geojson()

                # Fallback 1: provider-filtered DB rows (if provider column exists).
                if not entities and provider_column_exists:
                    res = await db.execute(
                        text(
                            """
                            SELECT basin_id AS entity_id, name, code
                            FROM geo.basin
                            WHERE UPPER(COALESCE(provider, 'ABH')) = 'DGM'
                            ORDER BY name
                            """
                        )
                    )
                    entities = [dict(r) for r in res.mappings().all()]

                # Fallback 2: generic DB list (last resort).
                if not entities:
                    res = await db.execute(text("SELECT basin_id AS entity_id, name, code FROM geo.basin ORDER BY name"))
                    entities = [dict(r) for r in res.mappings().all()]
            else:
                if provider_column_exists:
                    res = await db.execute(
                        text(
                            """
                            SELECT basin_id AS entity_id, name, code
                            FROM geo.basin
                            WHERE UPPER(COALESCE(provider, 'ABH')) = 'ABH'
                            ORDER BY name
                            """
                        )
                    )
                    entities = [dict(r) for r in res.mappings().all()]
                else:
                    res = await db.execute(text("SELECT basin_id AS entity_id, name, code FROM geo.basin ORDER BY name"))
                    entities = [dict(r) for r in res.mappings().all()]
        else:
            res = await db.execute(text("SELECT station_id AS entity_id, name, code FROM geo.station ORDER BY name"))
            entities = [dict(r) for r in res.mappings().all()]
    except:
        if is_basin_template:
            if normalized_basin_shape == "DGM":
                entities = _load_dgm_basins_from_geojson()
            if not entities:
                entities = [{"entity_id": "xxx", "name": "Bassin 1", "code": "B1"}]
        else:
            entities = [{"entity_id": "xxx", "name": "Station 1", "code": "S1"}]

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donnees"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    info_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

    # Info row
    shape_info = f" | Shape: {normalized_basin_shape}" if is_basin_template else ""
    ws["A1"] = f"Variable: {variable_label} ({variable_unit}) | Source: {source_label} ({normalized_source}){shape_info}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].fill = info_fill
    ws.merge_cells(f"A1:{get_column_letter(len(entities) + 1)}1")

    ws["A2"] = (
        f"Format: timestamp ISO8601 (YYYY-MM-DDTHH:MM:SS), valeurs numeriques par colonne {entity_label_singular} "
        f"(code en entete) | suffixe source: _{_source_suffix(normalized_source)}"
    )
    ws["A2"].fill = info_fill
    ws.merge_cells(f"A2:{get_column_letter(len(entities) + 1)}2")

    # Header row
    ws["A3"] = "timestamp"
    ws["A3"].font = header_font
    ws["A3"].fill = header_fill
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22

    for i, entity in enumerate(entities):
        col = get_column_letter(i + 2)
        cell = ws[f"{col}3"]
        entity_code = entity.get("code") or entity.get("name")
        cell.value = _append_source_suffix(entity_code, normalized_source)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[col].width = 16

    ws.row_dimensions[3].height = 35

    # Example rows
    base_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(3):
        dt = base_date - timedelta(days=i)
        row_num = 4 + i
        ws[f"A{row_num}"] = dt.strftime("%Y-%m-%dT%H:%M:%S")
        for j in range(len(entities)):
            ws[f"{get_column_letter(j + 2)}{row_num}"] = ""

    # Info sheet
    ws_info = wb.create_sheet("Bassins" if is_basin_template else "Stations")
    ws_info["A1"] = "Code"
    ws_info["B1"] = "Nom"
    for i, entity in enumerate(entities):
        entity_code = entity.get("code") or ""
        ws_info[f"A{i+2}"] = _append_source_suffix(entity_code, normalized_source) if entity_code else ""
        ws_info[f"B{i+2}"] = entity.get("name") or ""

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    shape_suffix = f"_{normalized_basin_shape.lower()}" if is_basin_template else ""
    filename = f"template_multi_{entity_label_plural}{shape_suffix}_{variable_code or 'variable'}_{normalized_source.lower()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/templates/multi-bassin")
async def get_template_multi_bassin(
    variable_code: Optional[str] = Query(None),
    source_code: Optional[str] = Query("OBS"),
    basin_shape: Optional[str] = Query("ABH"),
    db: AsyncSession = Depends(get_db)
):
    """Generate Excel template with columns = bassins (for one variable)."""
    return await get_template_multi_station(
        variable_code=variable_code,
        entity_type="bassins",
        source_code=source_code,
        basin_shape=basin_shape,
        db=db,
    )


@router.head("/templates/multi-bassin", include_in_schema=False)
async def head_template_multi_bassin():
    """Allow lightweight HEAD checks without triggering a 405."""
    return Response(status_code=200)


# --- SHP UPLOAD ---
@router.post("/shp/upload")
async def upload_shp(
    file: UploadFile = File(...), 
    dry_run: bool = Form(True),
    entity_type: str = Form("stations"),
    replace_mode: bool = Form(False),
    column_mapping: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    """
    Upload a SHP file or ZIP containing SHP, SHX, DBF, etc.
    Parse with GeoPandas.
    If dry_run=True, return GeoJSON preview.
    If dry_run=False, commit to DB (Insert/Update or Replace).
    """
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            if file.filename.lower().endswith('.zip'):
                zip_path = os.path.join(tmpdir, file.filename)
                with open(zip_path, "wb") as buffer:
                    content = await file.read()
                    buffer.write(content)
                
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(tmpdir)
                
                shp_file = None
                for root, dirs, files in os.walk(tmpdir):
                    for f in files:
                        if f.endswith(".shp"):
                            shp_file = os.path.join(root, f)
                            break
                
                if not shp_file:
                    raise HTTPException(status_code=400, detail="No .shp file found in ZIP")
            
            elif file.filename.lower().endswith('.shp'):
                shp_path = os.path.join(tmpdir, file.filename)
                with open(shp_path, "wb") as buffer:
                    content = await file.read()
                    buffer.write(content)
                shp_file = shp_path
            
            else:
                raise HTTPException(status_code=400, detail="File must be .shp or .zip containing shapefile")

            import geopandas as gpd
            gdf = gpd.read_file(shp_file)
            
            if gdf.crs and gdf.crs.to_string() != "EPSG:4326":
                gdf = gdf.to_crs("EPSG:4326")

            geojson_str = gdf.to_json()
            geojson_data = json.loads(geojson_str)
            
            if dry_run:
                return {
                    "status": "success",
                    "message": f"Parsed {len(gdf)} features from {os.path.basename(shp_file)}",
                    "message": f"Parsed {len(gdf)} features from {os.path.basename(shp_file)}",
                    "preview": geojson_data,
                    "columns": list(gdf.columns.astype(str)),
                    "debug_version": "v3"
                }
            
            target_table = "geo.station" if entity_type == "stations" else "geo.basin"
            if entity_type not in ["stations", "bassins", "basins"]:
                 raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")
            if entity_type in ["bassins", "basins"]:
                target_table = "geo.basin"

            gdf.columns = [c.lower() for c in gdf.columns]
            
            # Parse column mapping if provided
            mapping = {}
            if column_mapping:
                try:
                    mapping = json.loads(column_mapping)
                except:
                    pass

            code_col = (mapping.get("code") or next((c for c in gdf.columns if c in ['code', 'id', 'identifiant']), None))
            if code_col: code_col = code_col.lower()

            name_col = (mapping.get("name") or next((c for c in gdf.columns if c in ['name', 'nom', 'label']), None))
            if name_col: name_col = name_col.lower()

            # Use mapped type col, or auto-detect
            # Prioritize specific names over generic 'type'
            type_col = (mapping.get("type") or next((c for c in gdf.columns if c in ['station_type', 'type_station', 'type_statt', 'genre', 'type']), None))
            if type_col: type_col = type_col.lower()
            
            # Check for forced type
            forced_type = mapping.get("force_type")



# ... (imports remain)

# --- Helper for type mapping ---
            def get_station_type(row):
                # Legacy mapping for safety
                LEGACY_MAP = {
                    "pluviometrique": "Poste PluviomÃ©trique",
                    "poste_mesure": "Station hydrologique", 
                    "barrage": "Barrage",
                    "result_point": "point resultats",
                    "limnimetrique": "Station hydrologique",
                    "hydrologique": "Station hydrologique",
                    "poste pluviomÃ©trique": "Poste PluviomÃ©trique",
                    "station hydrologique": "Station hydrologique",
                    "point resultats": "point resultats"
                }

                import unicodedata
                
                if forced_type:
                    # Normalize input to NFC to ensure 'Ã©' is consistent
                    ft_norm = unicodedata.normalize('NFC', forced_type)
                    lower_forced_type = ft_norm.lower().strip()
                    
                    if "pluvio" in lower_forced_type: return "Poste PluviomÃ©trique"
                    if "barrage" in lower_forced_type: return "Barrage"
                    if "result" in lower_forced_type: return "point resultats"
                    if "hydrologique" in lower_forced_type or "station" in lower_forced_type: return "Station hydrologique"
                    
                    # Direct mapping fallback
                    return LEGACY_MAP.get(lower_forced_type, ft_norm)

                if not type_col: return "Station hydrologique"
                
                raw_val = str(row[type_col]).strip()
                val = raw_val.lower()
                
                print(f"DEBUG TYPE MAPPING: Col={type_col} | Raw='{raw_val}' | Lower='{val}'")

                if "pluvio" in val: return "Poste PluviomÃ©trique"
                if "barrage" in val: return "Barrage"
                if "hydrologique" in val or "limni" in val or "hydro" in val or "poste" in val: return "Station hydrologique"
                if "result" in val: return "point resultats"
                
                # print(f"DEBUG: Defaulting to Station hydrologique for val='{val}'")
                return "Station hydrologique"

            
            if not code_col:
                 failed_cols = list(gdf.columns)
                 raise HTTPException(status_code=400, detail=f"SHP must have a 'code', 'id' or 'identifiant' column. Found: {failed_cols}")
            
            # Basin column detection
            basin_col = (mapping.get("basin") or next((c for c in gdf.columns if c in ['basin', 'bassin', 'bassin_versant', 'bv']), None))
            if basin_col: basin_col = basin_col.lower()

            try:
                if replace_mode:
                    if target_table == "geo.station":
                        try:
                            await db.execute(text("DELETE FROM ts.measurement WHERE station_id IN (SELECT station_id FROM geo.station)"))
                        except Exception as e:
                            print(f"Warning cleaning measurements: {e}")
                    
                    await db.execute(text(f"TRUNCATE TABLE {target_table} CASCADE"))
                    
                    for idx, row in gdf.iterrows():
                        geom_wkt = row.geometry.wkt
                        code = str(row[code_col])
                        name = str(row[name_col]) if name_col else code
                        st_type = get_station_type(row)
                        
                        # Resolve Basin ID if column exists
                        basin_id = None
                        if basin_col and row.get(basin_col):
                            b_val = str(row[basin_col]).strip()
                            # Try to find basin by code or name
                            res_b = await db.execute(text("SELECT basin_id FROM geo.basin WHERE code = :v OR name = :v"), {"v": b_val})
                            b_row = res_b.first()
                            if b_row: basin_id = b_row.basin_id

                        stmt = text(f"""
                            INSERT INTO {target_table} (name, code, station_type, basin_id, geom)
                            VALUES (:name, :code, :st_type, :basin_id, ST_SetSRID(ST_GeomFromText(:geom), 4326))
                        """)
                        # print(f"DEBUG SQL REPLACE: st_type={st_type} basin_id={basin_id} code={code}")
                        await db.execute(stmt, {"name": name, "code": code, "st_type": st_type, "basin_id": basin_id, "geom": geom_wkt})

                        
                else:
                    for idx, row in gdf.iterrows():
                        geom_wkt = row.geometry.wkt
                        code = str(row[code_col])
                        name = str(row[name_col]) if name_col else code
                        st_type = get_station_type(row)
                        
                        # Resolve Basin ID if column exists
                        basin_id = None
                        if basin_col and row.get(basin_col):
                            b_val = str(row[basin_col]).strip()
                            res_b = await db.execute(text("SELECT basin_id FROM geo.basin WHERE code = :v OR name = :v"), {"v": b_val})
                            b_row = res_b.first()
                            if b_row: basin_id = b_row.basin_id
                        
                        check = await db.execute(text(f"SELECT 1 FROM {target_table} WHERE code = :code"), {"code": code})
                        exists = check.first()
                        
                        if exists:
                            # Update basin_id only if resolved
                            params = {"name": name, "code": code, "st_type": st_type, "geom": geom_wkt}
                            set_actions = "name = :name, station_type = :st_type, geom = ST_SetSRID(ST_GeomFromText(:geom), 4326)"
                            
                            if basin_id:
                                set_actions += ", basin_id = :basin_id"
                                params["basin_id"] = basin_id

                            stmt = text(f"""
                                UPDATE {target_table} 
                                SET {set_actions}
                                WHERE code = :code
                            """)
                            await db.execute(stmt, params)
                        else:
                            stmt = text(f"""
                                INSERT INTO {target_table} (name, code, station_type, basin_id, geom)
                                VALUES (:name, :code, :st_type, :basin_id, ST_SetSRID(ST_GeomFromText(:geom), 4326))
                            """)
                            # print(f"DEBUG SQL APPEND: st_type={st_type} basin_id={basin_id} code={code}")
                            await db.execute(stmt, {"name": name, "code": code, "st_type": st_type, "basin_id": basin_id, "geom": geom_wkt})

                
                await db.commit()
                return {"status": "success", "message": f"Successfully committed {len(gdf)} features to {target_table}"}

            except Exception as e:
                await db.rollback()
                raise e

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SHP Processing Error: {str(e)}")



