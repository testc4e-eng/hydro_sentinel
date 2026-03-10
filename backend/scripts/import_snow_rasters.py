from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
import argparse
import csv
import os
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import unquote, urlparse
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import create_engine, text


ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
SCHEMA_SQL = ROOT / "app" / "db" / "sebou_monitoring_schema.sql"
DEFAULT_DATASET_DIR = ROOT.parent / "Neige_Sebou_12-11-25_04-03-26"
DEFAULT_RASTER_DIRNAME = "4_SNOW_BINARY_SEBOU"
RASTER_PATTERN = "*.tif"
RASTER_DATE_PATTERN = re.compile(r"\.A(?P<year>\d{4})(?P<doy>\d{3})\.")


@dataclass(frozen=True)
class DbConn:
    host: str
    port: int
    user: str
    password: str
    database: str


@dataclass(frozen=True)
class SnowStatsRow:
    target_date: date
    snow_area_km2: Optional[float]
    snow_percentage: Optional[float]


def _read_env_file(path: Path) -> Dict[str, str]:
    values: Dict[str, str] = {}
    if not path.exists():
        return values
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip("\"'")
    return values


def _sync_db_url(url: str) -> str:
    if url.startswith("postgresql+asyncpg://"):
        return url.replace("postgresql+asyncpg://", "postgresql://", 1)
    if url.startswith("postgresql+psycopg://"):
        return url.replace("postgresql+psycopg://", "postgresql://", 1)
    return url


def _parse_db_url(url: str) -> DbConn:
    parsed = urlparse(_sync_db_url(url))
    if parsed.scheme not in {"postgresql", "postgres"}:
        raise RuntimeError(f"Unsupported DATABASE_URL scheme: {parsed.scheme}")
    if not parsed.hostname or not parsed.username or parsed.password is None or not parsed.path:
        raise RuntimeError("DATABASE_URL is incomplete (host/user/password/database required).")

    return DbConn(
        host=parsed.hostname,
        port=parsed.port or 5432,
        user=unquote(parsed.username),
        password=unquote(parsed.password),
        database=parsed.path.lstrip("/"),
    )


def _normalize_label(value: str) -> str:
    raw = value.strip().replace("²", "2")
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.lower()


def _find_stats_columns(headers: Iterable[str]) -> Tuple[str, str, str]:
    date_col: Optional[str] = None
    snow_area_col: Optional[str] = None
    snow_pct_col: Optional[str] = None

    for header in headers:
        normalized = _normalize_label(header)
        if date_col is None and "date" in normalized:
            date_col = header
        if snow_area_col is None and "superficie_neige" in normalized:
            snow_area_col = header
        if snow_pct_col is None and "pourcentage_neige" in normalized:
            snow_pct_col = header

    if not date_col or not snow_area_col or not snow_pct_col:
        raise RuntimeError(
            "Colonnes stats introuvables. Colonnes attendues proches de: "
            "'Date', 'Superficie_neige_km2', 'Pourcentage_neige_AOI'."
        )

    return date_col, snow_area_col, snow_pct_col


def _parse_date_value(raw_value: object) -> date:
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    value = str(raw_value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise RuntimeError(f"Format de date non supporte: {value}")


def _parse_optional_float(raw_value: object) -> Optional[float]:
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    normalized = str(raw_value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def _load_stats_from_xlsx(path: Path) -> Dict[date, SnowStatsRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header_values = next(rows)
    except StopIteration as exc:
        raise RuntimeError(f"Fichier stats vide: {path}") from exc

    headers = [str(value).strip() if value is not None else "" for value in header_values]
    date_col, snow_area_col, snow_pct_col = _find_stats_columns(headers)
    index_map = {header: idx for idx, header in enumerate(headers)}

    parsed: Dict[date, SnowStatsRow] = {}
    for values in rows:
        if not values:
            continue
        raw_date = values[index_map[date_col]] if index_map[date_col] < len(values) else None
        if raw_date is None or str(raw_date).strip() == "":
            continue

        target_date = _parse_date_value(raw_date)
        snow_area = _parse_optional_float(values[index_map[snow_area_col]] if index_map[snow_area_col] < len(values) else None)
        snow_pct = _parse_optional_float(values[index_map[snow_pct_col]] if index_map[snow_pct_col] < len(values) else None)
        parsed[target_date] = SnowStatsRow(
            target_date=target_date,
            snow_area_km2=snow_area,
            snow_percentage=snow_pct,
        )

    workbook.close()
    return parsed


def _load_stats_from_csv(path: Path) -> Dict[date, SnowStatsRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None:
            raise RuntimeError(f"CSV sans en-tete: {path}")

        date_col, snow_area_col, snow_pct_col = _find_stats_columns(reader.fieldnames)
        parsed: Dict[date, SnowStatsRow] = {}
        for row in reader:
            raw_date = row.get(date_col)
            if raw_date is None or str(raw_date).strip() == "":
                continue
            target_date = _parse_date_value(raw_date)
            parsed[target_date] = SnowStatsRow(
                target_date=target_date,
                snow_area_km2=_parse_optional_float(row.get(snow_area_col)),
                snow_percentage=_parse_optional_float(row.get(snow_pct_col)),
            )
        return parsed


def _find_stats_file(dataset_dir: Path, explicit_path: Optional[Path]) -> Path:
    if explicit_path:
        if not explicit_path.exists():
            raise FileNotFoundError(f"Fichier stats introuvable: {explicit_path}")
        return explicit_path

    candidates = sorted(dataset_dir.glob("*.xlsx")) + sorted(dataset_dir.glob("*.csv"))
    if not candidates:
        raise FileNotFoundError(
            f"Aucun fichier stats .xlsx/.csv trouve dans {dataset_dir}. "
            "Fournissez --stats-file si necessaire."
        )
    return candidates[0]


def _load_stats_table(stats_file: Path) -> Dict[date, SnowStatsRow]:
    suffix = stats_file.suffix.lower()
    if suffix == ".xlsx":
        return _load_stats_from_xlsx(stats_file)
    if suffix == ".csv":
        return _load_stats_from_csv(stats_file)
    raise RuntimeError(f"Format stats non supporte: {stats_file.suffix}")


def _parse_raster_date(raster_path: Path) -> date:
    match = RASTER_DATE_PATTERN.search(raster_path.name)
    if not match:
        raise RuntimeError(
            f"Date introuvable dans le nom de raster: {raster_path.name}. "
            "Format attendu contenant '.AYYYYDDD.'"
        )
    year = int(match.group("year"))
    day_of_year = int(match.group("doy"))
    return date(year, 1, 1) + timedelta(days=day_of_year - 1)


def _collect_polygon_geoms(geometry, acc: List[object]) -> None:
    geom_type = getattr(geometry, "geom_type", "")
    if geom_type == "Polygon":
        acc.append(geometry)
        return
    if geom_type == "MultiPolygon":
        acc.extend(list(geometry.geoms))
        return
    for child in getattr(geometry, "geoms", []):
        _collect_polygon_geoms(child, acc)


def _infer_source_epsg(src, raster_path: Path) -> Optional[int]:
    if src.crs is not None:
        epsg = src.crs.to_epsg()
        if epsg is not None:
            return int(epsg)

        crs_text = str(src.crs).upper()
        if "32630" in crs_text:
            return 32630
        if "32629" in crs_text:
            return 32629

    file_name = raster_path.name.upper()
    if "UTM30N" in file_name:
        return 32630
    if "UTM29N" in file_name:
        return 32629
    return None


def _extract_snow_geometry_wkt(raster_path: Path) -> Optional[Tuple[str, int]]:
    try:
        import numpy as np
        import rasterio
        from rasterio.features import shapes
        from shapely.geometry import MultiPolygon, shape
        from shapely.ops import unary_union
    except ImportError as exc:
        raise RuntimeError(
            "Dependances manquantes pour lire les rasters TIFF. "
            "Installez au minimum rasterio, numpy, shapely."
        ) from exc

    with rasterio.open(raster_path) as src:
        band = src.read(1)
        nodata = src.nodata

        valid_mask = np.ones(band.shape, dtype=bool)
        if nodata is not None:
            valid_mask &= band != nodata
        positive_mask = valid_mask & (band > 0)

        if not positive_mask.any():
            return None

        geoms = []
        for geom_mapping, pixel_value in shapes(band, mask=positive_mask, transform=src.transform):
            try:
                value = float(pixel_value)
            except Exception:
                continue
            if value <= 0:
                continue
            geom = shape(geom_mapping)
            if geom.is_empty:
                continue
            geoms.append(geom)

        if not geoms:
            return None

        merged = unary_union(geoms)
        polygons: List[object] = []
        _collect_polygon_geoms(merged, polygons)
        if not polygons:
            return None

        multipolygon = MultiPolygon(polygons)
        if multipolygon.is_empty:
            return None

        epsg = _infer_source_epsg(src, raster_path)
        if epsg is None:
            raise RuntimeError(f"EPSG introuvable dans le raster: {raster_path}")

        return multipolygon.wkt, int(epsg)


def _upsert_daily_statistics(conn, row: SnowStatsRow, stats_file_name: str) -> None:
    conn.execute(
        text(
            """
            INSERT INTO sebou.daily_statistics (
                date,
                snow_area_km2,
                snow_percentage,
                quality_score,
                data_sources,
                processing_time_seconds
            )
            VALUES (
                :target_date,
                :snow_area_km2,
                :snow_percentage,
                :quality_score,
                :data_sources,
                :processing_time_seconds
            )
            ON CONFLICT (date) DO UPDATE SET
                snow_area_km2 = EXCLUDED.snow_area_km2,
                snow_percentage = EXCLUDED.snow_percentage,
                quality_score = EXCLUDED.quality_score,
                data_sources = EXCLUDED.data_sources,
                processing_time_seconds = EXCLUDED.processing_time_seconds
            """
        ),
        {
            "target_date": row.target_date,
            "snow_area_km2": row.snow_area_km2,
            "snow_percentage": row.snow_percentage,
            "quality_score": 90,
            "data_sources": [f"snow_stats:{stats_file_name}"],
            "processing_time_seconds": 75,
        },
    )


def _insert_snow_extent(
    conn,
    target_date: date,
    geometry_wkt: str,
    source_srid: int,
    snow_area_km2: Optional[float],
    sensor: str,
) -> int:
    result = conn.execute(
        text(
            """
            INSERT INTO sebou.snow_extents (
                date,
                area_km2,
                detection_confidence,
                sensor,
                geom
            )
            SELECT
                :target_date,
                COALESCE(:snow_area_km2, ST_Area(geom_out) / 1000000.0),
                :confidence,
                :sensor,
                geom_out
            FROM (
                SELECT
                    ST_Multi(
                        ST_CollectionExtract(
                            ST_MakeValid(
                                ST_Transform(
                                    ST_SetSRID(ST_GeomFromText(:geometry_wkt), :source_srid),
                                    32629
                                )
                            ),
                            3
                        )
                    ) AS geom_out
            ) AS prepared
            WHERE geom_out IS NOT NULL
              AND NOT ST_IsEmpty(geom_out)
            """
        ),
        {
            "target_date": target_date,
            "snow_area_km2": snow_area_km2,
            "confidence": 0.85,
            "sensor": sensor,
            "geometry_wkt": geometry_wkt,
            "source_srid": source_srid,
        },
    )
    return int(result.rowcount or 0)


def _upsert_quality_report(conn, target_date: date, sensor: str) -> None:
    conn.execute(
        text(
            """
            DELETE FROM sebou.quality_reports
            WHERE processing_date = :target_date
              AND sensor = :sensor
            """
        ),
        {"target_date": target_date, "sensor": sensor},
    )
    conn.execute(
        text(
            """
            INSERT INTO sebou.quality_reports (
                processing_date,
                sensor,
                quality_flags,
                validation_score
            )
            VALUES (
                :target_date,
                :sensor,
                ARRAY[]::text[],
                :validation_score
            )
            """
        ),
        {
            "target_date": target_date,
            "sensor": sensor,
            "validation_score": 90,
        },
    )


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Importe les rasters neige (TIFF binaire) + stats (xlsx/csv) vers "
            "sebou.daily_statistics et sebou.snow_extents."
        )
    )
    parser.add_argument(
        "--dataset-dir",
        default=str(DEFAULT_DATASET_DIR),
        help=f"Dossier dataset neige (defaut: {DEFAULT_DATASET_DIR})",
    )
    parser.add_argument(
        "--raster-subdir",
        default=DEFAULT_RASTER_DIRNAME,
        help=f"Sous-dossier des rasters (defaut: {DEFAULT_RASTER_DIRNAME})",
    )
    parser.add_argument(
        "--stats-file",
        default=None,
        help="Chemin explicite vers le fichier stats (.xlsx ou .csv).",
    )
    parser.add_argument(
        "--sensor",
        default="MODIS",
        help="Nom capteur pour quality_reports/snow_extents (defaut: MODIS).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Analyse seulement (pas d'ecriture DB).",
    )
    return parser


def main() -> None:
    args = _build_arg_parser().parse_args()

    dataset_dir = Path(args.dataset_dir).resolve()
    raster_dir = dataset_dir / args.raster_subdir
    stats_file = _find_stats_file(dataset_dir, Path(args.stats_file).resolve() if args.stats_file else None)

    if not dataset_dir.exists():
        raise FileNotFoundError(f"Dataset introuvable: {dataset_dir}")
    if not raster_dir.exists():
        raise FileNotFoundError(f"Dossier rasters introuvable: {raster_dir}")

    stats_by_date = _load_stats_table(stats_file)
    raster_files = sorted(raster_dir.glob(RASTER_PATTERN))
    if not raster_files:
        raise FileNotFoundError(f"Aucun raster trouve dans {raster_dir} ({RASTER_PATTERN})")

    raster_dates = [_parse_raster_date(path) for path in raster_files]
    imported_dates = sorted(set(stats_by_date.keys()) | set(raster_dates))
    if not imported_dates:
        raise RuntimeError("Aucune date a importer.")

    print(f"[INFO] Dataset: {dataset_dir}")
    print(f"[INFO] Stats: {stats_file.name} ({len(stats_by_date)} lignes)")
    print(f"[INFO] Rasters: {len(raster_files)} fichiers")
    print(f"[INFO] Periode: {imported_dates[0]} -> {imported_dates[-1]}")

    if args.dry_run:
        print("[DRY-RUN] Aucune ecriture en base effectuee.")
        return

    env = _read_env_file(ENV_PATH)
    db_url = env.get("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL introuvable dans backend/.env")

    conn_info = _parse_db_url(db_url)
    os.environ["PGHOST"] = conn_info.host
    engine = create_engine(_sync_db_url(db_url).replace("postgresql://", "postgresql+psycopg://"), future=True)

    min_date = imported_dates[0]
    max_date = imported_dates[-1]
    inserted_extents = 0
    skipped_empty = 0

    with engine.begin() as conn:
        conn.execute(text(SCHEMA_SQL.read_text(encoding="utf-8")))

        conn.execute(
            text(
                """
                DELETE FROM sebou.snow_extents
                WHERE date BETWEEN :min_date AND :max_date
                """
            ),
            {"min_date": min_date, "max_date": max_date},
        )

        for stats_row in stats_by_date.values():
            _upsert_daily_statistics(conn, row=stats_row, stats_file_name=stats_file.name)
            _upsert_quality_report(conn, target_date=stats_row.target_date, sensor=args.sensor)

        for raster_path in raster_files:
            target_date = _parse_raster_date(raster_path)
            stats_row = stats_by_date.get(target_date)
            declared_area = stats_row.snow_area_km2 if stats_row else None

            if declared_area is not None and declared_area <= 0:
                skipped_empty += 1
                continue

            geometry = _extract_snow_geometry_wkt(raster_path)
            if geometry is None:
                skipped_empty += 1
                continue

            geometry_wkt, source_srid = geometry
            inserted_extents += _insert_snow_extent(
                conn,
                target_date=target_date,
                geometry_wkt=geometry_wkt,
                source_srid=source_srid,
                snow_area_km2=declared_area,
                sensor=args.sensor,
            )

    print("[DONE] Import neige termine.")
    print(f"[DONE] Snow extents inseres: {inserted_extents}")
    print(f"[DONE] Rasters sans neige/geom (ignores): {skipped_empty}")


if __name__ == "__main__":
    main()
